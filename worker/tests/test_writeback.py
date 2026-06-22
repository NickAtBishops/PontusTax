import datetime as dt

from openpyxl import load_workbook

from pontus_tax.canonical import (
    DELINQUENT, NEEDS_REVIEW, PAID, HIGH, LOW, MEDIUM,
    AccountRecord, RowOutcome,
)
from pontus_tax.intake import parse_workbook
from pontus_tax.writeback import output_filename, write_output

RUN_DATE = dt.date(2026, 6, 9)


def _outcomes():
    return {
        "s00_r0003": RowOutcome(
            row_key="s00_r0003", sheet_name="Florida Prop Tax", row_number=3,
            row_status=PAID, confidence=HIGH,
            status_note=(
                "Paid in full $4,974.48 on 12/29/2025 "
                "(Receipt N12292025P015431, paid by Robert Machin Jr)"
            ),
            write_date_paid="12/29/2025",
            write_receipt="N12292025P015431",
        ),
        "s00_r0004": RowOutcome(
            row_key="s00_r0004", sheet_name="Florida Prop Tax", row_number=4,
            row_status=NEEDS_REVIEW, confidence=LOW,
            status_note="NEEDS REVIEW — account not found",
            write_receipt="SHOULD-NOT-BE-WRITTEN",
        ),
        "s00_r0005": RowOutcome(
            row_key="s00_r0005", sheet_name="Florida Prop Tax", row_number=5,
            row_status=DELINQUENT, confidence=MEDIUM,
            status_note="DELINQUENT — $123,456.78 owed as of 6/9/2026",
            write_amount_due=123456.78,
            write_receipt="XYZ-OVERWRITE-ATTEMPT",
        ),
    }


def test_writeback_protections_and_notes_column(florida_workbook, tmp_path):
    # New layout (§10): no per-run column append. The notes sentence
    # OVERWRITES the rightmost existing month-update column — here "May
    # 2026 Update" at column V — every run. The dedicated Amount Due
    # column has been removed entirely; live owed values reach the
    # workbook only via the existing canonical amounts column (or, on
    # installment grids with multiple amount columns, only via the
    # notes sentence).
    with open(florida_workbook, "rb") as fh:
        original_bytes = fh.read()

    intake = parse_workbook(florida_workbook)
    out_path = str(tmp_path / output_filename("Property Taxes- Florida.xlsx", RUN_DATE))
    headers = write_output(intake, _outcomes(), RUN_DATE, out_path)

    # §10.4 — the original upload is never modified
    with open(florida_workbook, "rb") as fh:
        assert fh.read() == original_bytes
    assert out_path.endswith("Property Taxes- Florida — checked 2026-06-09.xlsx")
    # Return shape now exposes only the notes-column header (no "amount").
    assert headers["Florida Prop Tax"] == {"status": "May 2026 Update"}
    assert "amount" not in headers["Florida Prop Tax"]

    ws = load_workbook(out_path)["Florida Prop Tax"]

    # Notes land in V (May 2026 Update) — overwritten cleanly per run.
    assert ws["V2"].value == "May 2026 Update"  # header itself is untouched
    assert ws["V3"].value.startswith("Paid in full $4,974.48")
    assert ws["V4"].value == "NEEDS REVIEW — account not found"
    assert ws["V5"].value.startswith("DELINQUENT")

    # April 2026 Update (U) is a HISTORICAL artifact — left alone.
    # The fixture's row 3 has "Paud in full" pre-typed in U; that survives.
    assert ws["U3"].value == "Paud in full"

    # Verified payment details still write into their canonical columns
    # (R = Date Paid, S = Paid Confirmation #).
    assert ws["R3"].value == dt.datetime(2025, 12, 29)
    assert ws["S3"].value == "N12292025P015431"

    # LOW/NEEDS_REVIEW results never reach data cells (§7).
    assert ws["S4"].value is None
    # No silent erasure: existing receipt survives a conflicting scrape.
    assert ws["S5"].value == "KEEP-ME-123"

    # Formulas everywhere are untouched.
    assert ws["Q3"].value == "=N3+O3"
    assert ws["N7"].value == "=SUM(N3:N6)"
    # Live owed amount NOT written into installment grids (multiple
    # amounts columns) — it rides the notes sentence instead.
    assert ws["N5"].value == 120802.06


def test_no_new_column_appended(florida_workbook, ptax_master_workbook, tmp_path):
    # Two back-to-back runs on the same output must not grow the sheet.
    # Florida case: month-update column already exists → no new column
    # ever. PTAX case: 'Last run notes' is created exactly once on run
    # one and reused on run two.
    for label, fixture in (
        ("florida", florida_workbook),
        ("ptax", ptax_master_workbook),
    ):
        first = str(tmp_path / f"{label}-run1.xlsx")
        second = str(tmp_path / f"{label}-run2.xlsx")
        intake = parse_workbook(fixture)
        write_output(intake, _outcomes(), RUN_DATE, first)
        # Second run loads the run-1 output as its input — the realistic
        # repeated-run case.
        intake2 = parse_workbook(first)
        write_output(intake2, _outcomes(), RUN_DATE, second)
        ws1 = load_workbook(first).worksheets[0]
        ws2 = load_workbook(second).worksheets[0]
        assert ws1.max_column == ws2.max_column, (
            f"{label}: max_column grew between back-to-back runs "
            f"({ws1.max_column} → {ws2.max_column})"
        )


def test_main_sheet_layout_unchanged(florida_workbook, tmp_path):
    # Under the fixed-layout model (§10) the sheet width never grows on
    # legacy workbooks: notes overwrite the rightmost existing month-
    # update column, no per-run column append.
    intake = parse_workbook(florida_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, _outcomes(), RUN_DATE, out_path)

    in_wb = load_workbook(florida_workbook)
    out_wb = load_workbook(out_path)

    # Sheet count + names unchanged — guards against accidental Authorities
    # / metadata sheets ever sneaking back in.
    assert out_wb.sheetnames == in_wb.sheetnames == ["Florida Prop Tax"]

    in_ws, out_ws = in_wb["Florida Prop Tax"], out_wb["Florida Prop Tax"]
    in_header_max = max(c for c in range(1, in_ws.max_column + 1)
                        if in_ws.cell(row=2, column=c).value is not None)
    out_header_max = max(c for c in range(1, out_ws.max_column + 1)
                         if out_ws.cell(row=2, column=c).value is not None)
    assert in_header_max == 23
    assert out_header_max == 23


# ---------------------------------------------------------------------------
# PTAX_Master template — fixed-cell write path (S–V).
# ---------------------------------------------------------------------------


def _ptax_paid_putnam_outcome():
    # Synthesizes the canonical PAID-with-discount Putnam outcome targeting
    # PTAX_Master row 4 (the template's first data row): ultimate=0, paid
    # $4,974.48 on 2025-12-29, next deadline 2026-03-31.
    rec = AccountRecord(
        account_searched="09-05-24-005954-056-00",
        status=PAID,
        amount_due=0.0,
        ultimate_payment_due=0.0,
        amount_paid=4974.48,
        date_paid="2025-12-29",
        next_due_date="2026-03-31",
        confidence=HIGH,
    )
    return RowOutcome(
        row_key="s00_r0004",
        sheet_name="Property Tax",
        row_number=4,
        accounts=[rec],
        row_status=PAID,
        confidence=HIGH,
        status_note="PAID in full $4,974.48 on 2025-12-29",
        write_ultimate_payment_due=0.0,
        write_payment_date="2025-12-29",
        write_payment_amount=4974.48,
        write_next_due_date="2026-03-31",
    )


def test_structured_cells_written(ptax_master_workbook, tmp_path):
    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, {"s00_r0004": _ptax_paid_putnam_outcome()}, RUN_DATE, out_path)
    ws = load_workbook(out_path)["Property Tax"]

    # The four structured cells carry typed Excel values (number for $, real
    # datetime for dates), not strings.
    assert ws["S4"].value == 0.0
    assert ws["T4"].value == dt.datetime(2025, 12, 29)
    assert ws["U4"].value == 4974.48
    assert ws["V4"].value == dt.datetime(2026, 3, 31)


def test_template_fill_preserved(ptax_master_workbook, tmp_path):
    # Cream fill (FFFFF8DC) is the visual cue the analyst uses to spot the
    # auto-populated columns. Overwriting the value must not strip it.
    before = load_workbook(ptax_master_workbook)["Property Tax"]
    assert before["S4"].fill.fgColor.rgb == "FFFFF8DC"
    assert before["T4"].fill.fgColor.rgb == "FFFFF8DC"
    assert before["U4"].fill.fgColor.rgb == "FFFFF8DC"
    assert before["V4"].fill.fgColor.rgb == "FFFFF8DC"

    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, {"s00_r0004": _ptax_paid_putnam_outcome()}, RUN_DATE, out_path)
    after = load_workbook(out_path)["Property Tax"]

    for col in ("S", "T", "U", "V"):
        assert after[f"{col}4"].fill.fgColor.rgb == "FFFFF8DC", (
            f"{col}4 lost the cream fill after writeback"
        )


def test_unprocessed_rows_are_marked_not_skipped(florida_workbook, tmp_path):
    intake = parse_workbook(florida_workbook)
    outcomes = _outcomes()
    del outcomes["s00_r0005"]  # pretend the run died before row 5
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, outcomes, RUN_DATE, out_path)
    ws = load_workbook(out_path)["Florida Prop Tax"]
    # Notes column for the legacy Florida fixture is V (May 2026 Update).
    assert ws["V5"].value == "NOT CHECKED — run ended before this row"
