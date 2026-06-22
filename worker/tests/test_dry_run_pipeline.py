"""End-to-end through the real orchestrator with --dry-run semantics:
intake → per-row loop → write-back → summary, no portals contacted."""

import asyncio
import datetime as dt
import glob
import os

from openpyxl import load_workbook

from pontus_tax.canonical import (
    HIGH, PAID, RowOutcome, aggregate_status,
)
from pontus_tax.config import Config
from pontus_tax.orchestrator import RowProcessor, execute_run
from pontus_tax.store import LocalStore
from pontus_tax.validate import build_account_record, build_row_note
from pontus_tax.verify import MatchVerdict


def test_dry_run_full_pipeline(florida_workbook):
    store = LocalStore(florida_workbook)
    cfg = Config()
    summary = asyncio.run(execute_run(store, cfg, dry_run=True))

    assert summary["status_counts"] == {"NEEDS_REVIEW": 4}
    assert summary["status_column_header"].endswith("Update")
    assert summary["amount_column_header"].startswith("Amount Due")
    assert len(summary["review_rows"]) == 4

    out_dir = os.path.dirname(store.output_local)
    outputs = glob.glob(os.path.join(out_dir, "* — checked *.xlsx"))
    assert outputs, "checked workbook must be produced even on a dry run"
    ws = load_workbook(outputs[0])["Florida Prop Tax"]
    for row in (3, 4, 5, 6):
        assert "dry run" in str(ws.cell(row=row, column=25).value)
        assert ws.cell(row=row, column=24).value is None  # nothing invented

    summaries = glob.glob(
        os.path.join(os.path.dirname(florida_workbook), "* — summary *.json")
    )
    assert summaries, "run summary JSON must be written in local mode"


def test_payment_fields_propagate_end_to_end(florida_workbook, monkeypatch):
    # Verify the new structured payment fields survive the full
    # build_account_record → RowOutcome → store → write-back chain. The
    # other rows fall through the existing dry-run NEEDS_REVIEW path; only
    # the Putnam row gets a synthesized "successful" extraction.
    putnam_extraction = {
        "page_outcome": "account_found",
        "amount_due_now": 0,
        "ultimate_payment_due": 0,
        "payment_date": "2025-12-29",
        "payment_amount": 4974.48,
        "includes_delinquency": False,
        "owner_on_page": "Pontus EHC Putnam LLC",
        "situs_address_on_page": "800 REID ST",
        "parcel_or_account_on_page": "R09-009-009",
        "final_url": "https://putnam.county-taxes.net/public/R09-009-009",
    }
    verdict = MatchVerdict(True, "account + owner", owner_mismatch=False,
                           confidence_hint=HIGH)
    today = dt.date(2026, 6, 9)

    original_process = RowProcessor.process

    async def patched_process(self, job):
        if job.row.row_number != 6:
            return await original_process(self, job)
        rec = build_account_record("R09-009-009", putnam_extraction, verdict)
        outcome = RowOutcome(
            row_key=job.key,
            sheet_name=job.sheet.name,
            row_number=job.row.row_number,
            accounts=[rec],
            row_status=aggregate_status([rec.status]),
            confidence=rec.confidence,
            evidence=rec.evidence or "",
            status_note=build_row_note(
                [rec], aggregate_status([rec.status]), today,
            ),
            write_date_paid=rec.date_paid,
            write_amount_due=0.0 if rec.status == PAID else rec.amount_due,
        )
        return outcome

    monkeypatch.setattr(RowProcessor, "process", patched_process)

    store = LocalStore(florida_workbook)
    cfg = Config()
    summary = asyncio.run(execute_run(store, cfg, dry_run=True))

    # The non-Putnam rows still flow through the dry-run NEEDS_REVIEW path,
    # so the run summary counts 3 NEEDS_REVIEW + 1 PAID.
    assert summary["status_counts"].get("PAID") == 1
    assert summary["status_counts"].get("NEEDS_REVIEW") == 3

    # Saved outcome retains the new payment fields after passing through
    # the store boundary (LocalStore stores RowOutcome directly; the
    # FirestoreStore counterpart serializes via asdict — covered by the
    # roundtrip test in test_validate).
    saved = store.outcomes["s00_r0006"].accounts[0]
    assert saved.ultimate_payment_due == 0
    assert saved.amount_paid == 4974.48
    assert saved.date_paid == "2025-12-29"

    # The new wording reaches the workbook's status column.
    outputs = glob.glob(
        os.path.join(os.path.dirname(store.output_local), "* — checked *.xlsx")
    )
    ws = load_workbook(outputs[0])["Florida Prop Tax"]
    assert ws.cell(row=6, column=25).value == (
        "PAID in full $4,974.48 on 2025-12-29"
    )
    # Verified-paid → $0.00 in the dedicated Amount Due column.
    assert ws.cell(row=6, column=24).value == 0.0
    # The relocated SUM row remains intact at row 7.
    assert ws["N7"].value == "=SUM(N3:N6)"
