"""Shared fixture: a synthetic Florida-shaped workbook reproducing the
hostile patterns of the real tracker (CLAUDE.md §9): two stacked header
rows with merged group cells, PID vs Account #, installment grids typed by
data, a per-row Total FORMULA column, a totals row with SUM formulas,
a three-account cell, monthly update columns, and a hyperlink URL cell.
"""

from __future__ import annotations

import datetime as dt
import os
import sys

import pytest
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


@pytest.fixture
def florida_workbook(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Florida Prop Tax"

    # Row 1 — group header row (merged ranges)
    ws.merge_cells("A1:F1")
    ws["A1"] = "Property Information"
    ws.merge_cells("K1:M1")
    ws["K1"] = "Due Dates"
    ws.merge_cells("N1:P1")
    ws["N1"] = "Amounts"

    # Row 2 — the real field header row
    headers = {
        "A": "Property Address", "B": "City", "C": "St", "D": "Zip",
        "E": "County", "F": "Owner Entity", "G": "PID", "H": "Account #",
        "I": "Year of Assessment", "J": "# Installments",
        "K": "#1", "L": "#2", "M": "Early Bird",
        "N": "#1", "O": "#2", "P": "Early Bird",
        "Q": "Total", "R": "Date Paid", "S": "Paid Confirmation #",
        "T": "Tenant Responsible", "U": "April 2026 Update",
        "V": "May 2026 Update", "W": "Website",
    }
    for col, text in headers.items():
        ws[f"{col}2"] = text

    # Row 3 — Palm Beach (deep link via hyperlink cell)
    ws["A3"] = "950 Evernia St"
    ws["B3"] = "West Palm Beach"
    ws["C3"] = "FL"
    ws["D3"] = "33401"
    ws["E3"] = "Palm Beach"
    ws["F3"] = "EHC Palm Beach LLC"
    ws["G3"] = "VP4601"
    ws["H3"] = "#74-43-43-21-01-043-0050"
    ws["I3"] = 2025
    ws["J3"] = 2
    ws["K3"] = dt.datetime(2025, 11, 30)
    ws["L3"] = dt.datetime(2026, 3, 31)
    ws["N3"] = 1000.00
    ws["O3"] = 2000.00
    ws["Q3"] = "=N3+O3"
    ws["U3"] = "Paud in full"  # the sheet's own typo, kept as-is
    ws["W3"] = "Palm Beach portal"
    ws["W3"].hyperlink = (
        "https://pbctax.publicaccessnow.com/PropertyTax/Account.aspx"
        "?p=74-43-43-21-01-043-0050&a=1418360"
    )

    # Row 4 — Pinellas, THREE accounts in one cell (§5.6)
    ws["A4"] = "2180 49th St N"
    ws["B4"] = "St Petersburg"
    ws["C4"] = "FL"
    ws["D4"] = "33710"
    ws["E4"] = "Pinellas"
    ws["F4"] = "Pontus EHC Pinellas LLC"
    ws["G4"] = "VP4602"
    ws["H4"] = "#T815151/#T813795/#R444958"
    ws["I4"] = 2025
    ws["N4"] = 500.00
    ws["O4"] = 700.00
    ws["Q4"] = "=N4+O4"
    ws["W4"] = "https://pinellas.county-taxes.net/public"

    # Row 5 — Broward, NO url (portal must be discovered), existing receipt
    ws["A5"] = "1800 NW 49th St"
    ws["B5"] = "Fort Lauderdale"
    ws["C5"] = "FL"
    ws["D5"] = "33309"
    ws["E5"] = "Broward"
    ws["F5"] = "Pontus EHC Broward LLC"
    ws["G5"] = "VP4603"
    ws["H5"] = "504209AB0120"
    ws["I5"] = 2025
    ws["N5"] = 120802.06
    ws["Q5"] = "=N5+O5"
    ws["S5"] = "KEEP-ME-123"  # pre-existing receipt must never be erased

    # Row 6 — totals row: SUM formulas only, no identity → not a data row
    ws["N6"] = "=SUM(N3:N5)"
    ws["O6"] = "=SUM(O3:O5)"
    ws["Q6"] = "=SUM(Q3:Q5)"

    path = tmp_path / "Property Taxes- Florida.xlsx"
    wb.save(path)
    return str(path)
