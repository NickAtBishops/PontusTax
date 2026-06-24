"""Tests for excel_writer.write_quarterly_values against the real tracker.

We use samples/Corporate_Financials_and_P_Ls.xlsx as the fixture rather
than a synthesized one so the tests catch any drift between the worker's
assumptions and the real spreadsheet structure (e.g., if the Pinnacle row
moves, or the Q1 26 column header typo gets fixed in the source).
"""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Make the worker module importable when running pytest from the worker/
# directory without an installed package.
sys.path.insert(0, str(Path(__file__).parent))

from excel_writer import (  # noqa: E402
    WriteRequest,
    WritebackRefusedError,
    write_quarterly_values,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
MASTER_XLSX = REPO_ROOT / "samples" / "Corporate_Financials_and_P_Ls.xlsx"


@pytest.fixture
def master_bytes() -> bytes:
    return MASTER_XLSX.read_bytes()


def pinnacle_q1_2026_request(master_bytes: bytes, **overrides) -> WriteRequest:
    """The canonical happy-path request: Pinnacle row 14, Q1 2026, with
    the AI3 'Q4 26' typo alternate enabled."""
    defaults = dict(
        xlsx_bytes=master_bytes,
        sheet_name="Corp Financials ",  # trailing space is real
        row=14,
        expected_tenant_substring="Pinnacle",
        sales_col=17,
        ebitda_col=35,
        sales_value=51700,
        ebitda_value=7461,
        sales_header_expected="Q1 26",
        ebitda_header_expected="Q1 26",
        ebitda_header_alternate="Q4 26",
        header_row=3,
    )
    defaults.update(overrides)
    return WriteRequest(**defaults)


# ---------- happy path ----------


def test_happy_path_writes_both_cells(master_bytes: bytes):
    resp = write_quarterly_values(pinnacle_q1_2026_request(master_bytes))
    wb = load_workbook(BytesIO(resp.xlsx_bytes), keep_links=True)
    ws = wb["Corp Financials "]
    assert ws.cell(row=14, column=17).value == 51700
    assert ws.cell(row=14, column=35).value == 7461


def test_happy_path_preserves_existing_sales_formulas(master_bytes: bytes):
    """The Q2-Q4 2025 Sales cells use formulas in the source. We must not
    touch them."""
    resp = write_quarterly_values(pinnacle_q1_2026_request(master_bytes))
    wb = load_workbook(BytesIO(resp.xlsx_bytes), keep_links=True)
    ws = wb["Corp Financials "]
    # The exact formulas recorded during Phase 1 inspection.
    assert ws.cell(row=14, column=13).value == 25333.628  # Q1 25 hardcoded
    assert ws.cell(row=14, column=14).value == "=62450.083-M14"
    assert ws.cell(row=14, column=15).value == "=109177-N14-M14"
    assert ws.cell(row=14, column=16).value == "=157831-O14-N14-M14"


def test_happy_path_preserves_existing_ebitda_formulas(master_bytes: bytes):
    """All four 2025 EBITDA cells (AE-AH 14) are formulas. Must survive."""
    resp = write_quarterly_values(pinnacle_q1_2026_request(master_bytes))
    wb = load_workbook(BytesIO(resp.xlsx_bytes), keep_links=True)
    ws = wb["Corp Financials "]
    assert ws.cell(row=14, column=31).value == "=4252.541+430.088"  # Q1 25
    assert ws.cell(row=14, column=32).value == "=9209+54.5+962.6-AE14"
    assert ws.cell(row=14, column=33).value == "=15070+1033+1684-AF14-AE14"
    assert (
        ws.cell(row=14, column=34).value
        == "=18842+2291+1306-AG14-AF14-AE14"
    )


def test_happy_path_preserves_annual_and_ltm_formulas(master_bytes: bytes):
    """The 2025 annual Sales (T14) and LTM (V14) reference the quarterly
    cells we did NOT touch. They must still hold their original formulas
    so the new Q1 26 write doesn't accidentally flow into 2025 totals."""
    resp = write_quarterly_values(pinnacle_q1_2026_request(master_bytes))
    wb = load_workbook(BytesIO(resp.xlsx_bytes), keep_links=True)
    ws = wb["Corp Financials "]
    # 2025 sales annual = SUM of M..P
    assert ws.cell(row=14, column=20).value == "=SUM(M14:P14)"
    # LTM = same SUM (2025 LTM)
    assert ws.cell(row=14, column=22).value == "=SUM(M14:P14)"


def test_happy_path_reports_ebitda_header_typo_as_warning(master_bytes: bytes):
    resp = write_quarterly_values(pinnacle_q1_2026_request(master_bytes))
    assert len(resp.warnings) == 1
    assert "Q4 26" in resp.warnings[0]
    assert "typo" in resp.warnings[0].lower()


def test_other_tenants_rows_untouched(master_bytes: bytes):
    """Random sanity: every other row's column A (tenant name) is
    unchanged after the write."""
    original = load_workbook(BytesIO(master_bytes), keep_links=True)["Corp Financials "]
    resp = write_quarterly_values(pinnacle_q1_2026_request(master_bytes))
    new = load_workbook(BytesIO(resp.xlsx_bytes), keep_links=True)["Corp Financials "]
    for row in range(1, original.max_row + 1):
        assert new.cell(row=row, column=1).value == original.cell(row=row, column=1).value


def test_produced_xlsx_opens_without_openpyxl_warnings(master_bytes: bytes):
    """The structural cell-diff test passes, but Excel can still flash a
    'we found a problem with some content - do you want us to recover'
    dialog if the xlsx has subtle integrity issues (broken shared
    strings, malformed styles, dangling drawing references). openpyxl
    surfaces those via UserWarning during load. Treat any UserWarning
    on a clean re-load as a regression."""
    import warnings

    resp = write_quarterly_values(pinnacle_q1_2026_request(master_bytes))
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        load_workbook(BytesIO(resp.xlsx_bytes), keep_links=True)
    user_warnings = [
        w for w in caught if issubclass(w.category, UserWarning)
    ]
    assert user_warnings == [], (
        f"openpyxl raised {len(user_warnings)} UserWarning(s) when loading "
        f"the produced xlsx, which often correlates with Excel showing a "
        f"'we found a problem with some content' dialog. Warnings: "
        f"{[str(w.message) for w in user_warnings]}"
    )


def test_corp_financials_sheet_has_only_intended_changes(master_bytes: bytes):
    """Strict structural check: every cell of the Corp Financials sheet
    in the output is bit-identical to the original EXCEPT Q14 (Sales Q1
    26) and AI14 (EBITDA Q1 26).

    NB: this does NOT check the P&Ls sheet. openpyxl's float
    serialization perturbs the last 1-2 decimals of double-precision
    floats during a roundtrip even for cells we never touched (worst
    case observed: 4.66e-10 absolute drift on a $2.1M value, i.e. half
    a nanopenny). That noise is invisible in Excel because cells are
    formatted to 2 decimal places, but it shows up at the bit level
    on the P&Ls sheet which has many imported / computed floats.
    The Corp Financials sheet is mostly integer-thousands and formulas,
    so it stays bit-identical."""
    resp = write_quarterly_values(pinnacle_q1_2026_request(master_bytes))
    orig_sheet = load_workbook(BytesIO(master_bytes), keep_links=True)["Corp Financials "]
    new_sheet = load_workbook(BytesIO(resp.xlsx_bytes), keep_links=True)["Corp Financials "]
    diffs = []
    for row in range(1, max(orig_sheet.max_row, new_sheet.max_row) + 1):
        for col in range(1, max(orig_sheet.max_column, new_sheet.max_column) + 1):
            if orig_sheet.cell(row, col).value != new_sheet.cell(row, col).value:
                diffs.append((row, col))
    assert diffs == [(14, 17), (14, 35)], (
        "Unexpected diffs on Corp Financials beyond the two intended "
        f"cells: {diffs}"
    )


# ---------- precondition refusals ----------


def test_refuses_when_sheet_missing(master_bytes: bytes):
    with pytest.raises(WritebackRefusedError, match="not in workbook"):
        write_quarterly_values(
            pinnacle_q1_2026_request(master_bytes, sheet_name="Nope")
        )


def test_refuses_when_tenant_substring_mismatch(master_bytes: bytes):
    """Row 14 holds Pinnacle. If we say expect "Trulieve" there, the
    worker must refuse rather than write into Pinnacle's row."""
    with pytest.raises(WritebackRefusedError, match="does not contain"):
        write_quarterly_values(
            pinnacle_q1_2026_request(
                master_bytes, expected_tenant_substring="Trulieve"
            )
        )


def test_refuses_when_sales_header_mismatch(master_bytes: bytes):
    # Col 17 row 3 is "Q1 26". Telling the worker to expect "Q2 26" must
    # abort.
    with pytest.raises(WritebackRefusedError, match="Sales column header"):
        write_quarterly_values(
            pinnacle_q1_2026_request(master_bytes, sales_header_expected="Q2 26")
        )


def test_refuses_when_ebitda_header_mismatch_without_alternate(master_bytes: bytes):
    # AI3 reads "Q4 26" in the source. Without the alternate enabled,
    # the worker should refuse because the actual header doesn't match
    # "Q1 26".
    with pytest.raises(WritebackRefusedError, match="EBITDA column header"):
        write_quarterly_values(
            pinnacle_q1_2026_request(
                master_bytes, ebitda_header_alternate=None
            )
        )


def test_refuses_to_overwrite_a_formula(master_bytes: bytes):
    """Target the Q2 25 EBITDA cell (AF14 = col 32) which is a formula.
    Worker must refuse."""
    with pytest.raises(WritebackRefusedError, match="contains a formula"):
        write_quarterly_values(
            pinnacle_q1_2026_request(
                master_bytes,
                ebitda_col=32,
                ebitda_header_expected="Q2 25",
                ebitda_header_alternate=None,
            )
        )


def test_refuses_to_overwrite_a_hardcoded_value(master_bytes: bytes):
    """Target the Q1 25 Sales cell (M14 = col 13, value 25333.628).
    Not a formula, but already populated. Worker must refuse."""
    with pytest.raises(WritebackRefusedError, match="already holds the value"):
        write_quarterly_values(
            pinnacle_q1_2026_request(
                master_bytes,
                sales_col=13,
                sales_header_expected="Q1 25",
            )
        )


def test_refuses_when_target_row_a_is_empty_for_tenant(master_bytes: bytes):
    """Row 1 column A is "Key Info" (a header). Expect "Pinnacle" in
    column A and you'll be told it doesn't match — worker refuses."""
    with pytest.raises(WritebackRefusedError, match="does not contain"):
        write_quarterly_values(
            pinnacle_q1_2026_request(master_bytes, row=1)
        )
