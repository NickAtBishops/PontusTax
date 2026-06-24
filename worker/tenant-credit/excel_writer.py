"""Cell-level write-back into the Corp Financials tracker.

Pure function: bytes of the original xlsx and a typed request in, modified
xlsx bytes and a list of soft warnings out. Refuses to write rather than
risk corrupting the analyst's spreadsheet whenever the situation looks
even slightly off.

What "off" means here:
  - The named sheet isn't in the workbook.
  - The target row's column A doesn't contain the expected tenant name,
    so we're about to write into the wrong tenant.
  - The column header in the spec row (row 3 by default) doesn't match
    the quarter we were told to write. There's one allowlisted typo on
    the EBITDA Q1 26 column (AI3 = "Q4 26" in the source); other
    mismatches abort.
  - The target cell already holds a formula. Overwriting a formula
    destroys the analyst's work and there's no good reason to do it.
  - The target cell already holds a literal value. We refuse rather than
    silently re-stamp it; the analyst should clear it explicitly first
    if they really do mean to re-write.

The original xlsx in samples/ is treated as read-only by convention; this
module never writes to disk. The caller saves the returned bytes under a
new filename (with a timestamp) so the master copy is preserved.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook


@dataclass(frozen=True)
class WriteRequest:
    # The original xlsx bytes. Never mutated; openpyxl reads from a
    # BytesIO copy.
    xlsx_bytes: bytes

    # Exact sheet name including any trailing whitespace ("Corp
    # Financials " with the trailing space is a real example).
    sheet_name: str

    # 1-indexed row that the target tenant lives on. Pulled from the
    # tenant's TypeScript config on the caller side.
    row: int

    # We require column A of `row` to contain this substring,
    # case-insensitively. Guard against writing to the wrong tenant if
    # the spreadsheet's row order ever shifts.
    expected_tenant_substring: str

    # 1-indexed columns where Sales and EBITDA for the target quarter
    # live. The caller derives them from the tracker layout module.
    sales_col: int
    ebitda_col: int

    # The integer thousands values to write. Floats accepted but the
    # engine rounds first; we don't apply any additional rounding here.
    sales_value: float
    ebitda_value: float

    # The row-3 (or whatever header_row) label we expect to find above
    # the Sales and EBITDA columns for the requested quarter, e.g. "Q1
    # 26". A mismatch is a hard error because it almost certainly means
    # we're about to write into the wrong quarter.
    sales_header_expected: str
    ebitda_header_expected: str

    # If set, the EBITDA header column is allowed to read this value
    # instead of `ebitda_header_expected`. The known case: AI3 in the
    # actual tracker reads "Q4 26" instead of "Q1 26" due to a typo.
    # The worker writes anyway but reports the discrepancy in warnings.
    ebitda_header_alternate: Optional[str] = None

    # Row that carries quarter labels. Defaults to 3 to match the
    # Corp Financials sheet.
    header_row: int = 3


@dataclass
class WriteResponse:
    xlsx_bytes: bytes
    # Soft messages the caller should surface in the audit log or UI:
    # for example, "EBITDA header was the known typo".
    warnings: list[str] = field(default_factory=list)


class WritebackRefusedError(ValueError):
    """Raised when a precondition fails. The caller maps this to HTTP 422
    rather than 500: it's analyst-actionable, not a server bug."""


def write_quarterly_values(req: WriteRequest) -> WriteResponse:
    """Open the workbook, validate, write the two cells, return new bytes."""
    # keep_links=True so external workbook links survive the round trip.
    # data_only is left at its default False so we read formulas as
    # formulas (we'd otherwise see their last cached values instead of
    # the formula strings we need to inspect).
    wb = load_workbook(BytesIO(req.xlsx_bytes), keep_links=True)

    if req.sheet_name not in wb.sheetnames:
        raise WritebackRefusedError(
            f"Sheet '{req.sheet_name}' not in workbook. "
            f"Available sheets: {wb.sheetnames!r}."
        )
    ws = wb[req.sheet_name]

    warnings: list[str] = []

    # --- 1. Tenant identity check on column A of the target row. ---
    name_cell = ws.cell(row=req.row, column=1)
    name_text = "" if name_cell.value is None else str(name_cell.value)
    needle = req.expected_tenant_substring
    if needle.lower() not in name_text.lower():
        raise WritebackRefusedError(
            f"Row {req.row} column A is {name_text!r}, which does not "
            f"contain expected tenant substring {needle!r}. Refusing to "
            "write to a row that may belong to a different tenant."
        )

    # --- 2. Sales column header check. ---
    sales_header_cell = ws.cell(row=req.header_row, column=req.sales_col)
    sales_header = (
        "" if sales_header_cell.value is None else str(sales_header_cell.value)
    )
    if sales_header != req.sales_header_expected:
        raise WritebackRefusedError(
            f"Sales column header at row {req.header_row} col "
            f"{req.sales_col} is {sales_header!r}, expected "
            f"{req.sales_header_expected!r}. Refusing to write to a "
            "column that doesn't match the requested quarter."
        )

    # --- 3. EBITDA column header check (with allowlisted typo). ---
    ebitda_header_cell = ws.cell(row=req.header_row, column=req.ebitda_col)
    ebitda_header = (
        "" if ebitda_header_cell.value is None
        else str(ebitda_header_cell.value)
    )
    if ebitda_header == req.ebitda_header_expected:
        pass
    elif (
        req.ebitda_header_alternate is not None
        and ebitda_header == req.ebitda_header_alternate
    ):
        warnings.append(
            f"EBITDA column header at row {req.header_row} col "
            f"{req.ebitda_col} is {ebitda_header!r}, the known typo for "
            f"{req.ebitda_header_expected!r}. Writing anyway; consider "
            "fixing the header in the spreadsheet."
        )
    else:
        raise WritebackRefusedError(
            f"EBITDA column header at row {req.header_row} col "
            f"{req.ebitda_col} is {ebitda_header!r}, expected "
            f"{req.ebitda_header_expected!r}"
            + (
                f" (or alternate {req.ebitda_header_alternate!r})"
                if req.ebitda_header_alternate is not None
                else ""
            )
            + ". Refusing to write."
        )

    # --- 4. Target cells must be empty and must not be formulas. ---
    sales_cell = ws.cell(row=req.row, column=req.sales_col)
    ebitda_cell = ws.cell(row=req.row, column=req.ebitda_col)

    for label, cell, col in [
        ("Sales", sales_cell, req.sales_col),
        ("EBITDA", ebitda_cell, req.ebitda_col),
    ]:
        # openpyxl uses data_type 'f' for formula cells. Reading the
        # value of a formula cell returns the formula string (with
        # data_only=False), not its computed result.
        if cell.data_type == "f":
            raise WritebackRefusedError(
                f"{label} target cell (row {req.row}, col {col}) "
                f"contains a formula: {cell.value!r}. Refusing to "
                "overwrite a formula."
            )
        if cell.value is not None:
            raise WritebackRefusedError(
                f"{label} target cell (row {req.row}, col {col}) "
                f"already holds the value {cell.value!r}. Refusing to "
                "overwrite a non-empty cell; clear it first if you "
                "really intend to overwrite."
            )

    # --- 5. Write the values. ---
    sales_cell.value = req.sales_value
    ebitda_cell.value = req.ebitda_value

    # --- 6. Serialize back to bytes. The caller is responsible for
    # naming the saved file. ---
    out = BytesIO()
    wb.save(out)
    return WriteResponse(xlsx_bytes=out.getvalue(), warnings=warnings)
