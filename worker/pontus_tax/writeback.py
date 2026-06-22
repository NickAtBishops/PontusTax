"""Write-back — CLAUDE.md §10.

openpyxl, preserving structure: formatting, widths, merged headers,
hyperlinks, formulas. Writes ONLY into detected canonical columns plus one
NEW status column named after the workbook's own pattern. Formula cells are
never touched; a scraped blank/zero never erases a real value; no row is
silently skipped. The original upload is never modified — output is a copy.
"""

from __future__ import annotations

import datetime as dt
import logging
from copy import copy
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .canonical import NEEDS_REVIEW, PAID, UNREACHABLE, LOW, RowOutcome
from .intake import (
    NOTES_HEADER_DEFAULT, SheetIntake, WorkbookIntake, status_column_header,
)
from .validate import parse_date, parse_money

# Number formats used when overwriting PTAX_Master's S–V cells. Excel will
# round-trip the underlying datetime / number; the template already sets
# these formats on the sample rows, but we re-apply for sheets that don't.
_CURRENCY_FMT = '"$"#,##0.00'
_DATE_FMT = "yyyy-mm-dd"

# PTAX_Master reserves W..AJ for analyst-filled values and formulas:
#   W–Y    Jurisdiction Links              (analyst-filled)
#   Z–AC   BOV — Single Broker             (analyst-filled)
#   AD–AF  BOV — Median of Multiple        (analyst-filled)
#   AG     Actual assessment               (analyst-filled today; Stage 2 may scrape)
#   AH/AI/AJ  Variance ($) / (%) / Appeal flag — EXCEL FORMULAS the
#              variance chain depends on. Replacing them with hardcoded
#              values breaks the chain forever.
# The writeback must never touch any of these. The single exception is
# the resolved notes column (which may sit past V on fresh layouts that
# require creating "Last run notes" once at ws.max_column + 1).
_PROTECTED_RANGE_START = 23  # W (1-indexed)


class WritebackGuardError(RuntimeError):
    """Raised when write-back attempts to touch a column past V that is not
    the resolved notes column. Surfaces with the offending column letter
    and header text so the misconfigured caller is obvious."""


def _assert_writable_column(
    ws, sheet: SheetIntake, target_col: int, notes_col: int,
) -> None:
    if target_col < _PROTECTED_RANGE_START or target_col == notes_col:
        return
    header = ws.cell(row=sheet.header_row, column=target_col).value
    raise WritebackGuardError(
        f"refusing to write into protected column "
        f"{get_column_letter(target_col)} ({header!r}) — "
        f"columns W..AJ are analyst-filled or formula columns; the "
        f"writeback must only touch S–V or the resolved notes column"
    )

log = logging.getLogger("pontus_tax.writeback")

# Statuses whose data values are allowed into data cells (§7: LOW/NEEDS_REVIEW
# results reach the status column only).
_WRITABLE_STATUSES_EXCLUDED = {NEEDS_REVIEW, UNREACHABLE}


def _is_formula(cell) -> bool:
    return cell.data_type == "f" or (
        isinstance(cell.value, str) and str(cell.value).startswith("=")
    )


def _cell_empty(cell) -> bool:
    v = cell.value
    return v is None or (isinstance(v, str) and not v.strip())


def _values_equalish(existing: Any, new: Any) -> bool:
    if existing is None or new is None:
        return False
    e_money, n_money = parse_money(existing), parse_money(new)
    if e_money is not None and n_money is not None:
        return abs(e_money - n_money) < 0.01
    e_date, n_date = parse_date(existing), parse_date(new)
    if e_date is not None and n_date is not None:
        return e_date == n_date
    return str(existing).strip().lower() == str(new).strip().lower()


def _safe_write(ws, row: int, col: int, value: Any, number_format: str | None = None) -> bool:
    """One gate for every data-cell write: never a formula cell, never erase,
    never fight an existing different value (corrections ride the status
    column instead). Returns True when the cell was written."""
    cell = ws.cell(row=row, column=col)
    if _is_formula(cell):
        return False
    if value is None or (isinstance(value, str) and not value.strip()):
        return False  # §7 no silent erasure — blanks never overwrite
    if not _cell_empty(cell):
        return _values_equalish(cell.value, value)  # already right → fine
    cell.value = value
    if number_format:
        cell.number_format = number_format
    return True


def _write_structured_cell(ws, row: int, col: int, value: Any, number_format: str) -> bool:
    """Overwrite a fixed structured cell (S–V on PTAX_Master), preserving the
    template's per-cell fill (cream FFFFF8DC), font, and alignment.
    cell.number_format = ... can drop style attributes on some openpyxl
    versions, so save and restore explicitly. None / empty string never
    overwrites (§7: no silent erasure). The richer correction rules —
    no $0 over an existing nonzero, date-correction-with-note — layer on
    in a later commit; this helper handles the basic write."""
    cell = ws.cell(row=row, column=col)
    if _is_formula(cell):
        return False
    if value is None or (isinstance(value, str) and not value.strip()):
        return False
    saved_fill = copy(cell.fill)
    saved_font = copy(cell.font)
    saved_align = copy(cell.alignment)
    cell.value = value
    cell.number_format = number_format
    cell.fill = saved_fill
    cell.font = saved_font
    cell.alignment = saved_align
    return True


def _iso_to_datetime(iso: str | None) -> dt.datetime | None:
    """Parse an ISO-ish date string into datetime.datetime at midnight, the
    form openpyxl writes as a real Excel date value (not a string)."""
    d = parse_date(iso) if iso else None
    return dt.datetime(d.year, d.month, d.day) if d else None


def _collect_correction(cell, new_value: Any, header_label: str) -> str | None:
    """Compare an existing cell's value to a new value about to overwrite it.

    Returns a correction sentence when the portal contradicts a real
    prior value ("corrected <field> from <old> to <new> per portal
    receipt"). Returns None when the existing cell is empty, when both
    represent the same logical value (so an unchanged write looks like
    no-op), or when comparison isn't meaningful.
    """
    existing = cell.value
    if existing is None or (isinstance(existing, str) and not existing.strip()):
        return None

    def _fmt(v) -> str:
        if isinstance(v, dt.datetime):
            return v.date().isoformat()
        if isinstance(v, dt.date):
            return v.isoformat()
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    # Date-vs-date / date-vs-iso-string normalization.
    new_d = (
        new_value.date() if isinstance(new_value, dt.datetime)
        else new_value if isinstance(new_value, dt.date)
        else None
    )
    if new_d is not None:
        e_d = (
            existing.date() if isinstance(existing, dt.datetime)
            else existing if isinstance(existing, dt.date)
            else parse_date(existing) if isinstance(existing, str)
            else None
        )
        if e_d is not None and e_d == new_d:
            return None
        return (
            f"corrected {header_label} from {_fmt(existing)} to "
            f"{new_d.isoformat()} per portal receipt"
        )

    # Number-vs-number tolerance (cents-level).
    if isinstance(new_value, (int, float)) and isinstance(existing, (int, float)):
        if abs(float(existing) - float(new_value)) < 0.005:
            return None
        return (
            f"corrected {header_label} from {_fmt(existing)} to "
            f"{_fmt(new_value)} per portal receipt"
        )

    # Mixed types or strings — string-normalized equality.
    if _fmt(existing).lower() == _fmt(new_value).lower():
        return None
    return (
        f"corrected {header_label} from {_fmt(existing)} to "
        f"{_fmt(new_value)} per portal receipt"
    )


def _column_number_format(ws, col: int, data_start: int, data_end: int) -> str | None:
    for r in range(data_start, min(data_end, data_start + 40) + 1):
        cell = ws.cell(row=r, column=col)
        if cell.value is not None and cell.number_format != "General":
            return cell.number_format
    return None


def _last_used_column(ws, sheet: SheetIntake) -> int:
    last = 0
    rows_to_scan = [sheet.header_row]
    if sheet.group_row:
        rows_to_scan.append(sheet.group_row)
    rows_to_scan += [r.row_number for r in sheet.rows[:3]]
    for r in rows_to_scan:
        for c in range(ws.max_column, 0, -1):
            if ws.cell(row=r, column=c).value is not None:
                last = max(last, c)
                break
    return last or ws.max_column


def _resolve_notes_column(ws, sheet: SheetIntake) -> int:
    """Return the single column index the notes sentence is written into.

    Three-branch rule (§10):
    1. Rightmost existing '<Month> YYYY Update' column wins (older
       Florida-style workbooks; historical month-stamped columns to its
       left are left untouched as artifacts).
    2. Existing 'Last run notes' column from a prior run on the same
       workbook (lookup by header text, case-insensitive).
    3. Otherwise create 'Last run notes' exactly once, at the rightmost
       column position, with a reasonable width.
    """
    if sheet.update_columns:
        return sheet.update_columns[-1].index
    target_name = NOTES_HEADER_DEFAULT.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=sheet.header_row, column=c).value
        if v is not None and str(v).strip().lower() == target_name:
            return c
    new_col = _last_used_column(ws, sheet) + 1
    ws.cell(row=sheet.header_row, column=new_col).value = NOTES_HEADER_DEFAULT
    ws.column_dimensions[get_column_letter(new_col)].width = 52
    return new_col


def write_output(
    intake: WorkbookIntake,
    outcomes: dict[str, RowOutcome],
    run_date: dt.date,
    out_path: str,
) -> dict[str, dict[str, str]]:
    """Produce the checked copy. Returns
    {sheet_name: {"status": <notes column header>}}.

    The new fixed-layout model (§10) does NOT append per-run columns.
    The status sentence overwrites a single notes cell whose column is
    resolved by `_resolve_notes_column`: the rightmost pre-existing
    month-update column on legacy workbooks, or a 'Last run notes'
    column created exactly once on first run for fresh layouts.

    `run_date` is retained as the as-of date the calling note-builder
    may use when composing the sentence; the writeback itself no longer
    derives a per-run column name from it."""
    wb = load_workbook(intake.path, data_only=False)
    headers_added: dict[str, dict[str, str]] = {}

    for s_idx, sheet in enumerate(intake.sheets):
        ws = wb[sheet.name]
        data_rows = [r.row_number for r in sheet.rows]
        data_start, data_end = min(data_rows), max(data_rows)

        # ---- Notes column: single, fixed-position, overwritten per run.
        notes_col = _resolve_notes_column(ws, sheet)
        notes_header = str(
            ws.cell(row=sheet.header_row, column=notes_col).value or ""
        ).strip() or status_column_header(sheet)
        headers_added[sheet.name] = {"status": notes_header}

        date_info = sheet.first_col("date_paid")
        conf_info = sheet.first_col("confirmation")
        assessed_info = sheet.first_col("assessed_value")
        amount_cols = sheet.columns.get("amounts", [])
        # PTAX_Master structured cells (S–V). Absent on legacy workbooks
        # like the older Florida sheet; present on the master template.
        ultimate_info = sheet.first_col("ultimate_payment_due")
        paydate_info = sheet.first_col("payment_date")
        payamt_info = sheet.first_col("payment_amount")
        nextdue_info = sheet.first_col("next_due_date")
        date_fmt = (
            _column_number_format(ws, date_info.index, data_start, data_end)
            if date_info
            else None
        ) or "MM/DD/YYYY"

        # Resolve which legacy-canonical fields are safe to write to:
        # one of them (assessed_value) maps to AG on PTAX_Master via the
        # fuzzy synonym table — that's the analyst-filled "Actual
        # assessment" column and the writeback must skip it. The guard
        # exists for *any* would-be write past V; the pre-checks below
        # turn the legacy fields into clean no-ops when their detected
        # column happens to land in the protected range.
        def _writable(info) -> bool:
            return info is not None and (
                info.index < _PROTECTED_RANGE_START or info.index == notes_col
            )

        for row in sheet.rows:
            key = f"s{s_idx:02d}_r{row.row_number:04d}"
            outcome = outcomes.get(key)
            _assert_writable_column(ws, sheet, notes_col, notes_col)
            note_cell = ws.cell(row=row.row_number, column=notes_col)
            if outcome is None:
                # §10.3 — no row is ever silently skipped.
                note_cell.value = "NOT CHECKED — run ended before this row"
                continue

            allowed = (
                outcome.row_status not in _WRITABLE_STATUSES_EXCLUDED
                and outcome.confidence != LOW
            )
            # Corrections accumulate during the structured writes below
            # and append to the notes sentence at the end of this row.
            corrections: list[str] = []

            if allowed:
                if _writable(date_info) and outcome.write_date_paid:
                    d = parse_date(outcome.write_date_paid)
                    if d:
                        _assert_writable_column(ws, sheet, date_info.index, notes_col)
                        _safe_write(ws, row.row_number, date_info.index, d, date_fmt)
                if _writable(conf_info) and outcome.write_receipt:
                    _assert_writable_column(ws, sheet, conf_info.index, notes_col)
                    _safe_write(ws, row.row_number, conf_info.index, outcome.write_receipt)
                if _writable(assessed_info) and outcome.write_assessed_value is not None:
                    _assert_writable_column(ws, sheet, assessed_info.index, notes_col)
                    _safe_write(
                        ws, row.row_number, assessed_info.index,
                        outcome.write_assessed_value,
                    )
                if (
                    len(amount_cols) == 1
                    and outcome.write_amount_due is not None
                    and _writable(amount_cols[0])
                ):
                    _assert_writable_column(ws, sheet, amount_cols[0].index, notes_col)
                    _safe_write(
                        ws, row.row_number, amount_cols[0].index,
                        outcome.write_amount_due,
                    )

                # ---- PTAX_Master structured cells (S–V). Fixed-position
                # overwrites every run, preserving cream fill. Each write
                # collects a correction note when the portal contradicts a
                # real prior value, so analysts can see what changed and why.
                if ultimate_info and outcome.write_ultimate_payment_due is not None:
                    cell = ws.cell(row=row.row_number, column=ultimate_info.index)
                    c = _collect_correction(
                        cell, outcome.write_ultimate_payment_due, "ultimate payment due",
                    )
                    if c:
                        corrections.append(c)
                    _assert_writable_column(ws, sheet, ultimate_info.index, notes_col)
                    _write_structured_cell(
                        ws, row.row_number, ultimate_info.index,
                        outcome.write_ultimate_payment_due, _CURRENCY_FMT,
                    )
                if paydate_info and outcome.write_payment_date:
                    pd = _iso_to_datetime(outcome.write_payment_date)
                    if pd:
                        cell = ws.cell(row=row.row_number, column=paydate_info.index)
                        c = _collect_correction(cell, pd, "payment date")
                        if c:
                            corrections.append(c)
                        _assert_writable_column(ws, sheet, paydate_info.index, notes_col)
                        _write_structured_cell(
                            ws, row.row_number, paydate_info.index, pd, _DATE_FMT,
                        )
                if payamt_info and outcome.write_payment_amount is not None:
                    cell = ws.cell(row=row.row_number, column=payamt_info.index)
                    c = _collect_correction(
                        cell, outcome.write_payment_amount, "payment amount",
                    )
                    if c:
                        corrections.append(c)
                    _assert_writable_column(ws, sheet, payamt_info.index, notes_col)
                    _write_structured_cell(
                        ws, row.row_number, payamt_info.index,
                        outcome.write_payment_amount, _CURRENCY_FMT,
                    )
                if nextdue_info and outcome.write_next_due_date:
                    nd = _iso_to_datetime(outcome.write_next_due_date)
                    if nd:
                        cell = ws.cell(row=row.row_number, column=nextdue_info.index)
                        c = _collect_correction(cell, nd, "next due date")
                        if c:
                            corrections.append(c)
                        _assert_writable_column(ws, sheet, nextdue_info.index, notes_col)
                        _write_structured_cell(
                            ws, row.row_number, nextdue_info.index, nd, _DATE_FMT,
                        )

            # Write the notes cell LAST so any corrections collected during
            # the structured writes ride along on the same line analysts
            # already read for run status.
            base = outcome.status_note or "NOT CHECKED"
            note_cell.value = (
                f"{base} | " + " | ".join(corrections) if corrections else base
            )

    wb.save(out_path)
    log.info("wrote checked workbook: %s", out_path)
    return headers_added


def output_filename(input_name: str, run_date: dt.date) -> str:
    stem, dot, ext = input_name.rpartition(".")
    if not dot:
        stem, ext = input_name, "xlsx"
    return f"{stem} — checked {run_date.isoformat()}.{ext}"
