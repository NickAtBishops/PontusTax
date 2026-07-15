import datetime as dt

from pontus_tax.intake import _match_field, parse_workbook, status_column_header


def test_header_detection_two_stacked_rows(florida_workbook):
    intake = parse_workbook(florida_workbook)
    assert len(intake.sheets) == 1
    sheet = intake.sheets[0]
    assert sheet.header_row == 2
    assert sheet.group_row == 1


def test_column_mapping_by_header_not_letter(florida_workbook):
    sheet = parse_workbook(florida_workbook).sheets[0]
    m = sheet.mapping_doc()
    assert m["address"] == "A"
    assert m["city"] == "B"
    assert m["state"] == "C"
    assert m["zip"] == "D"
    assert m["county"] == "E"
    assert m["owner_entity"] == "F"
    assert m["internal_id"] == "G"      # PID is Pontus-internal, NOT the parcel
    assert m["account_number"] == "H"   # the real search key
    assert m["tax_year"] == "I"
    assert m["installments"] == "J"
    assert m["date_paid"] == "R"
    assert m["confirmation"] == "S"
    assert m["responsible_party"] == "T"
    assert m["website"] == "W"
    assert m["total"] == "Q"
    assert set(m["status_notes"].split(",")) == {"U", "V"}


def test_numbered_columns_typed_by_data(florida_workbook):
    sheet = parse_workbook(florida_workbook).sheets[0]
    m = sheet.mapping_doc()
    # K/L hold dates, M is an (empty) date-group column; N/O/P hold numbers
    assert "K" in m["due_dates"] and "L" in m["due_dates"]
    assert "N" in m["amounts"] and "O" in m["amounts"]
    assert "K" not in m["amounts"]


def test_formula_total_column_protected_but_not_sum_row_columns(florida_workbook):
    sheet = parse_workbook(florida_workbook).sheets[0]
    # Q has a formula in EVERY data row → protected.
    assert "Q" in sheet.protected_columns
    # N/O carry one SUM in the totals row only — per-cell guards handle that;
    # the columns themselves stay writable.
    assert "N" not in sheet.protected_columns
    assert "O" not in sheet.protected_columns


def test_data_rows_exclude_totals_row(florida_workbook):
    sheet = parse_workbook(florida_workbook).sheets[0]
    assert [r.row_number for r in sheet.rows] == [3, 4, 5, 6]


def test_url_from_hyperlink_and_multi_account_cell(florida_workbook):
    sheet = parse_workbook(florida_workbook).sheets[0]
    rows = {r.row_number: r for r in sheet.rows}
    assert rows[3].url.startswith("https://pbctax.publicaccessnow.com/")
    assert rows[4].url == "https://pinellas.county-taxes.net/public"
    assert rows[5].url is None
    assert [g.display for g in rows[4].accounts] == [
        "T815151", "T813795", "R444958",
    ]
    assert rows[3].tax_year == "2025"
    assert rows[5].confirmation_existing == "KEEP-ME-123"


def test_status_column_follows_workbook_pattern(florida_workbook):
    # Legacy Florida-style workbooks have month-update columns; the new
    # write-back model OVERWRITES the rightmost one every run instead of
    # appending a fresh per-run column.
    sheet = parse_workbook(florida_workbook).sheets[0]
    assert sheet.update_columns, "April/May 2026 Update columns should be detected"
    assert status_column_header(sheet) == "May 2026 Update"


def test_status_column_defaults_to_last_run_notes_on_fresh_template(
    ptax_master_workbook,
):
    # PTAX_Master has no month-update column → the writeback creates
    # 'Last run notes' once; the header function reports that name.
    sheet = parse_workbook(ptax_master_workbook).sheets[0]
    assert sheet.update_columns == []
    assert status_column_header(sheet) == "Last run notes"


# ---------------------------------------------------------------------------
# PTAX_Master template — fixed-cell layout (the new canonical workbook).
# ---------------------------------------------------------------------------


def test_ptax_master_header_detection(ptax_master_workbook):
    # README sheet has no detectable header row → skipped naturally.
    # Property Tax sheet (2026-06-25 reshuffle): row 2 is the System /
    # Pontus System Report / Model System Output band, row 3 is section
    # sub-labels (Location / Business Plan / Property Tax / Due Date),
    # row 4 is the field-header row, data starts at row 5. detect_header
    # picks row 4 (most field matches) and row 3 as the group row above it.
    intake = parse_workbook(ptax_master_workbook)
    assert [s.name for s in intake.sheets] == ["Property Tax"]
    sheet = intake.sheets[0]
    assert sheet.header_row == 4
    assert sheet.group_row == 3


def test_ptax_master_structured_field_mapping(ptax_master_workbook):
    # The three structured cells the checker writes to land at Y/Z/AB
    # under the "Model System Output" band. The two legacy cells
    # (ultimate_payment_due, next_due_date) were dropped from the
    # 2026-06-25 layout and are absent — first_col returns None.
    sheet = parse_workbook(ptax_master_workbook).sheets[0]
    assert sheet.first_col("payment_amount").letter == "Y"
    assert sheet.first_col("payment_date").letter == "Z"
    assert sheet.first_col("run_confidence").letter == "AB"
    assert sheet.first_col("ultimate_payment_due") is None
    assert sheet.first_col("next_due_date") is None


def test_ptax_master_original_tracker_columns_still_map(ptax_master_workbook):
    # The strict-header additions must not break the existing fuzzy mapping
    # for the A–R analyst-maintained columns the checker READS.
    m = parse_workbook(ptax_master_workbook).sheets[0].mapping_doc()
    assert m["address"] == "A"
    assert m["city"] == "B"
    assert m["state"] == "C"
    assert m["zip"] == "D"
    assert m["owner_entity"] == "E"
    assert m["internal_id"] == "F"     # PID
    assert m["account_number"] == "G"  # Account #
    assert m["installments"] == "H"
    assert m["confirmation"] == "N"
    assert m["responsible_party"] == "O"
    assert m["website"] == "R"
    # "Actual assessment" at AA fuzzy-matches the assessed_value bucket;
    # the writeback then refuses to touch it (it isn't in
    # _WRITABLE_FIELDS) — see test_writeback_guard_blocks_protected_columns.
    assert m["assessed_value"] == "AA"


def test_strict_match_beats_fuzzy_for_payment_amount():
    # "Payment amount" contains the word "amount" which the fuzzy table
    # would otherwise route to the loose `amounts` bucket. Strict wins.
    field, spec, tied = _match_field("Payment amount")
    assert field == "payment_amount"
    assert tied == []
    assert spec >= 1000


def test_strict_match_beats_fuzzy_for_next_due_date():
    # "Next due date" contains "due date" — the fuzzy table would otherwise
    # route to `due_dates`. Strict wins.
    field, _, tied = _match_field("Next due date")
    assert field == "next_due_date"
    assert tied == []


def test_strict_header_is_case_and_whitespace_tolerant():
    # Case + extra whitespace must still match strictly. Paraphrase must not.
    assert _match_field("ULTIMATE   PAYMENT  DUE")[0] == "ultimate_payment_due"
    assert _match_field("  Payment Date  ")[0] == "payment_date"
    # NOT tolerant of paraphrase — "Amount due now" should not become
    # ultimate_payment_due. (It falls through to a fuzzy match elsewhere
    # or returns None; here we just assert the strict path doesn't fire.)
    assert _match_field("Amount due now")[0] != "ultimate_payment_due"


# ---------------------------------------------------------------------------
# Jurisdiction Link columns (V/W/X) — checker must visit these too, not just
# "Website". The word "link" in these headers otherwise fuzzy-collides with
# the `website` synonym table; STRICT_HEADERS gives them their own fields.
# ---------------------------------------------------------------------------


def test_jurisdiction_link_headers_resolve_via_strict_match_not_fuzzy():
    assert _match_field("Jurisdiction link primary")[0] == "jurisdiction_link_primary"
    assert _match_field("Jurisdiction link secondary")[0] == "jurisdiction_link_secondary"
    assert _match_field("Jurisdiction link tertiary")[0] == "jurisdiction_link_tertiary"
    # Case/whitespace tolerant, like every other strict header.
    assert _match_field("  JURISDICTION   LINK  PRIMARY  ")[0] == (
        "jurisdiction_link_primary"
    )


def test_jurisdiction_link_columns_no_longer_ambiguous_on_real_template(
    ptax_master_workbook,
):
    from openpyxl import load_workbook

    wb = load_workbook(ptax_master_workbook)
    ws = wb["Property Tax"]
    # Row 5 is the template's sample data row; V/W/X already carry a URL
    # (V) and are blank (W, X) — fill W/X too so all three are exercised.
    ws["W5"] = "https://example-secondary.county-taxes.net/public"
    ws["X5"] = "https://example-tertiary.county-taxes.net/public"
    wb.save(ptax_master_workbook)

    intake = parse_workbook(ptax_master_workbook)
    sheet = intake.sheets[0]
    m = sheet.mapping_doc()

    assert m["jurisdiction_link_primary"] == "V"
    assert m["jurisdiction_link_secondary"] == "W"
    assert m["jurisdiction_link_tertiary"] == "X"
    # "website" must map ONLY to R — none of V/W/X collide into it anymore.
    assert m["website"] == "R"

    # No ambiguous-column log entries mentioning V, W, or X.
    for letter in ("V", "W", "X"):
        assert not any(
            f"column {letter} " in note for note in sheet.ambiguous
        ), sheet.ambiguous

    row5 = next(r for r in sheet.rows if r.row_number == 5)
    assert row5.jurisdiction_link_primary == (
        "https://county-taxes.net/fl-clay/property-tax/"
        "Y2xheTpyZWFsX2VzdGF0ZTpwYXJlbnRzOjM2OTQwMzc0LTM0N2MtMTFlZC05NzQ1LThlM2ZlZjFhN2Q3MQ=="
    )
    assert row5.jurisdiction_link_secondary == (
        "https://example-secondary.county-taxes.net/public"
    )
    assert row5.jurisdiction_link_tertiary == (
        "https://example-tertiary.county-taxes.net/public"
    )


def test_row_intake_exposes_jurisdiction_link_fields(florida_workbook):
    # Even on a workbook with none of these columns, RowIntake still
    # carries the three attributes (defaulting to None).
    sheet = parse_workbook(florida_workbook).sheets[0]
    row = sheet.rows[0]
    assert row.jurisdiction_link_primary is None
    assert row.jurisdiction_link_secondary is None
    assert row.jurisdiction_link_tertiary is None


def _clear_cell(ws, ref: str) -> None:
    """Blank a cell's value AND any hyperlink object — plain `ws[ref] = None`
    leaves a pre-existing hyperlink target in place (openpyxl stores it
    separately from `.value`), which would make `_extract_url` keep
    resolving a URL from a cell the test meant to empty out."""
    cell = ws[ref]
    cell.value = None
    cell.hyperlink = None


def test_check_urls_orders_dedupes_and_skips_blank_website(ptax_master_workbook):
    from openpyxl import load_workbook

    wb = load_workbook(ptax_master_workbook)
    ws = wb["Property Tax"]
    # Row 5: give Website (R) and Jurisdiction link primary (V) distinct
    # URLs (the template ships them mirrored), duplicate Website's URL
    # into secondary (W) to prove dedup, and give tertiary (X) a third,
    # distinct URL.
    website_url = "https://example-website.county-taxes.net/public"
    primary_url = "https://example-primary.county-taxes.net/public"
    tertiary_url = "https://example-tertiary.county-taxes.net/public"
    _clear_cell(ws, "R5")
    _clear_cell(ws, "V5")
    ws["R5"] = website_url
    ws["V5"] = primary_url
    ws["W5"] = website_url  # exact-string duplicate of Website
    ws["X5"] = tertiary_url
    wb.save(ptax_master_workbook)

    sheet = parse_workbook(ptax_master_workbook).sheets[0]
    row5 = next(r for r in sheet.rows if r.row_number == 5)
    urls = row5.check_urls()

    # Website + primary + tertiary; secondary is deduped away (same string
    # as Website, and Website is listed first, so Website's label wins).
    assert urls == [
        ("Website", website_url),
        ("Jurisdiction link primary", primary_url),
        ("Jurisdiction link tertiary", tertiary_url),
    ]


def test_check_urls_includes_jurisdiction_link_when_website_blank(
    ptax_master_workbook,
):
    from openpyxl import load_workbook

    wb = load_workbook(ptax_master_workbook)
    ws = wb["Property Tax"]
    _clear_cell(ws, "R5")  # blank out Website (clears its hyperlink too)
    _clear_cell(ws, "V5")  # blank out the template's default primary link
    ws["W5"] = "https://only-jurisdiction-link.county-taxes.net/public"
    wb.save(ptax_master_workbook)

    sheet = parse_workbook(ptax_master_workbook).sheets[0]
    row5 = next(r for r in sheet.rows if r.row_number == 5)
    assert row5.url is None
    urls = row5.check_urls()
    assert urls == [
        (
            "Jurisdiction link secondary",
            "https://only-jurisdiction-link.county-taxes.net/public",
        )
    ]
