import datetime as dt

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pontus_tax.canonical import (
    DELINQUENT, NEEDS_REVIEW, PAID, HIGH, LOW, MEDIUM,
    AccountRecord, RowOutcome,
)
from pontus_tax.intake import parse_workbook
from pontus_tax.writeback import (
    WritebackGuardError, _assert_writable_column,
    output_filename, write_output,
)

RUN_DATE = dt.date(2026, 6, 9)


def _outcomes():
    return {
        "s00_r0003": RowOutcome(
            row_key="s00_r0003", sheet_name="Florida Prop Tax", row_number=3,
            row_status=PAID, confidence=HIGH,
            status_note=(
                "Paid in full $4,974.48 on 12/29/2025 "
                "(Receipt N12292025P015431, paid by Robert Machin Jr)"
            ),
            write_date_paid="12/29/2025",
            write_receipt="N12292025P015431",
        ),
        "s00_r0004": RowOutcome(
            row_key="s00_r0004", sheet_name="Florida Prop Tax", row_number=4,
            row_status=NEEDS_REVIEW, confidence=LOW,
            status_note="NEEDS REVIEW — account not found",
            write_receipt="SHOULD-NOT-BE-WRITTEN",
        ),
        "s00_r0005": RowOutcome(
            row_key="s00_r0005", sheet_name="Florida Prop Tax", row_number=5,
            row_status=DELINQUENT, confidence=MEDIUM,
            status_note="DELINQUENT — $123,456.78 owed as of 6/9/2026",
            write_amount_due=123456.78,
            write_receipt="XYZ-OVERWRITE-ATTEMPT",
        ),
    }


def test_writeback_protections_and_notes_column(florida_workbook, tmp_path):
    # New layout (§10): no per-run column append. The notes sentence
    # OVERWRITES the rightmost existing month-update column — here "May
    # 2026 Update" at column V — every run. The dedicated Amount Due
    # column has been removed entirely; live owed values reach the
    # workbook only via the existing canonical amounts column (or, on
    # installment grids with multiple amount columns, only via the
    # notes sentence).
    with open(florida_workbook, "rb") as fh:
        original_bytes = fh.read()

    intake = parse_workbook(florida_workbook)
    out_path = str(tmp_path / output_filename("Property Taxes- Florida.xlsx", RUN_DATE))
    headers = write_output(intake, _outcomes(), RUN_DATE, out_path)

    # §10.4 — the original upload is never modified
    with open(florida_workbook, "rb") as fh:
        assert fh.read() == original_bytes
    assert out_path.endswith("Property Taxes- Florida — checked 2026-06-09.xlsx")
    # Return shape now exposes only the notes-column header (no "amount").
    assert headers["Florida Prop Tax"] == {"status": "May 2026 Update"}
    assert "amount" not in headers["Florida Prop Tax"]

    ws = load_workbook(out_path)["Florida Prop Tax"]

    # Notes land in V (May 2026 Update) — overwritten cleanly per run.
    assert ws["V2"].value == "May 2026 Update"  # header itself is untouched
    assert ws["V3"].value.startswith("Paid in full $4,974.48")
    assert ws["V4"].value == "NEEDS REVIEW — account not found"
    assert ws["V5"].value.startswith("DELINQUENT")

    # April 2026 Update (U) is a HISTORICAL artifact — left alone.
    # The fixture's row 3 has "Paud in full" pre-typed in U; that survives.
    assert ws["U3"].value == "Paud in full"

    # Verified payment details still write into their canonical columns
    # (R = Date Paid, S = Paid Confirmation #).
    assert ws["R3"].value == dt.datetime(2025, 12, 29)
    assert ws["S3"].value == "N12292025P015431"

    # LOW/NEEDS_REVIEW results never reach data cells (§7).
    assert ws["S4"].value is None
    # No silent erasure: existing receipt survives a conflicting scrape.
    assert ws["S5"].value == "KEEP-ME-123"

    # Formulas everywhere are untouched.
    assert ws["Q3"].value == "=N3+O3"
    assert ws["N7"].value == "=SUM(N3:N6)"
    # Live owed amount NOT written into installment grids (multiple
    # amounts columns) — it rides the notes sentence instead.
    assert ws["N5"].value == 120802.06


def test_no_new_column_appended(florida_workbook, ptax_master_workbook, tmp_path):
    # Two back-to-back runs on the same output must not grow the sheet.
    # Florida case: month-update column already exists → no new column
    # ever. PTAX case: 'Last run notes' is created exactly once on run
    # one and reused on run two.
    for label, fixture in (
        ("florida", florida_workbook),
        ("ptax", ptax_master_workbook),
    ):
        first = str(tmp_path / f"{label}-run1.xlsx")
        second = str(tmp_path / f"{label}-run2.xlsx")
        intake = parse_workbook(fixture)
        write_output(intake, _outcomes(), RUN_DATE, first)
        # Second run loads the run-1 output as its input — the realistic
        # repeated-run case.
        intake2 = parse_workbook(first)
        write_output(intake2, _outcomes(), RUN_DATE, second)
        ws1 = load_workbook(first).worksheets[0]
        ws2 = load_workbook(second).worksheets[0]
        assert ws1.max_column == ws2.max_column, (
            f"{label}: max_column grew between back-to-back runs "
            f"({ws1.max_column} → {ws2.max_column})"
        )


def test_main_sheet_layout_unchanged(florida_workbook, tmp_path):
    # Under the fixed-layout model (§10) the sheet width never grows on
    # legacy workbooks: notes overwrite the rightmost existing month-
    # update column, no per-run column append.
    intake = parse_workbook(florida_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, _outcomes(), RUN_DATE, out_path)

    in_wb = load_workbook(florida_workbook)
    out_wb = load_workbook(out_path)

    # Sheet count + names unchanged — guards against accidental Authorities
    # / metadata sheets ever sneaking back in.
    assert out_wb.sheetnames == in_wb.sheetnames == ["Florida Prop Tax"]

    in_ws, out_ws = in_wb["Florida Prop Tax"], out_wb["Florida Prop Tax"]
    in_header_max = max(c for c in range(1, in_ws.max_column + 1)
                        if in_ws.cell(row=2, column=c).value is not None)
    out_header_max = max(c for c in range(1, out_ws.max_column + 1)
                         if out_ws.cell(row=2, column=c).value is not None)
    assert in_header_max == 23
    assert out_header_max == 23


# ---------------------------------------------------------------------------
# PTAX_Master template — fixed-cell write path (S–V).
# ---------------------------------------------------------------------------


# PTAX_Master template (2026-06-25 reshuffle): data starts at row 5 (the
# group row at 2 and section sub-labels at 3 pushed the field-header row
# down to 4). The template's first data row is row 5 — Putnam.
PTAX_FIRST_DATA_ROW = 5


def _ptax_paid_putnam_outcome():
    # Synthesizes the canonical PAID-with-discount Putnam outcome targeting
    # PTAX_Master row 5 (the template's first data row): paid $4,974.48 on
    # 2025-12-29. Confidence HIGH rides into the AB column. Legacy
    # ultimate_payment_due / next_due_date are still set on the outcome
    # but the current template has no header for them, so those writes
    # are clean no-ops.
    rec = AccountRecord(
        account_searched="09-05-24-005954-056-00",
        status=PAID,
        amount_due=0.0,
        ultimate_payment_due=0.0,
        amount_paid=4974.48,
        date_paid="2025-12-29",
        next_due_date="2026-03-31",
        confidence=HIGH,
    )
    return RowOutcome(
        row_key=f"s00_r{PTAX_FIRST_DATA_ROW:04d}",
        sheet_name="Property Tax",
        row_number=PTAX_FIRST_DATA_ROW,
        accounts=[rec],
        row_status=PAID,
        confidence=HIGH,
        status_note="PAID in full $4,974.48 on 2025-12-29",
        write_ultimate_payment_due=0.0,
        write_payment_date="2025-12-29",
        write_payment_amount=4974.48,
        write_next_due_date="2026-03-31",
    )


def test_structured_cells_written(ptax_master_workbook, tmp_path):
    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(
        intake,
        {f"s00_r{PTAX_FIRST_DATA_ROW:04d}": _ptax_paid_putnam_outcome()},
        RUN_DATE, out_path,
    )
    ws = load_workbook(out_path)["Property Tax"]

    # Three Model System Output cells carry typed Excel values (number for
    # $, real datetime for dates, plain string for confidence).
    r = PTAX_FIRST_DATA_ROW
    assert ws[f"Y{r}"].value == 4974.48                          # Payment amount
    assert ws[f"Z{r}"].value == dt.datetime(2025, 12, 29)        # Payment date
    assert ws[f"AB{r}"].value == "HIGH"                          # Confidence
    # Analyst columns inside the same band must be preserved byte-for-byte.
    # Row 5 (Putnam) ships with Actual assessment = 3666000 in the template.
    assert ws[f"AA{r}"].value == 3666000


def test_template_fill_preserved(ptax_master_workbook, tmp_path):
    # Cream fill (FFFFF8DC) is the visual cue the analyst uses to spot the
    # auto-populated columns. Overwriting the value must not strip it.
    r = PTAX_FIRST_DATA_ROW
    before = load_workbook(ptax_master_workbook)["Property Tax"]
    for col in ("Y", "Z", "AB"):
        assert before[f"{col}{r}"].fill.fgColor.rgb == "FFFFF8DC", (
            f"template missing cream fill at {col}{r}"
        )

    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(
        intake,
        {f"s00_r{r:04d}": _ptax_paid_putnam_outcome()},
        RUN_DATE, out_path,
    )
    after = load_workbook(out_path)["Property Tax"]

    for col in ("Y", "Z", "AB"):
        assert after[f"{col}{r}"].fill.fgColor.rgb == "FFFFF8DC", (
            f"{col}{r} lost the cream fill after writeback"
        )


def _ptax_notes_col(out_path):
    ws = load_workbook(out_path)["Property Tax"]
    # Header row on PTAX_Master is row 4 (post-2026-06-25 layout).
    for c in range(1, ws.max_column + 1):
        v = ws.cell(4, c).value
        if v and str(v).strip().lower() == "last run notes":
            return c
        if v and "update" in str(v).lower():
            return c
    raise AssertionError("notes column not found")


def test_contradiction_routes_to_notes_only(ptax_master_workbook, tmp_path):
    # ultimate_payment_due > 0 with status PAID is the cross_check
    # contradiction case: cross_check_ultimate_due flips status to
    # NEEDS_REVIEW upstream, the orchestrator's gate leaves every
    # write_* field unset, and the writeback skips every structured
    # write. The contradiction sentence reaches the notes column with the
    # 'flagged for human review' marker so analysts see WHY.
    r = PTAX_FIRST_DATA_ROW
    record = AccountRecord(
        account_searched="09-05-24-005954-056-00",
        status=NEEDS_REVIEW,
        ultimate_payment_due=1234.0,
        confidence=LOW,
        evidence=(
            "matched by account + owner; contradiction: ultimate due "
            "$1,234.00 but reported PAID — flagged for human review"
        ),
    )
    outcome = RowOutcome(
        row_key=f"s00_r{r:04d}",
        sheet_name="Property Tax",
        row_number=r,
        accounts=[record],
        row_status=NEEDS_REVIEW,
        confidence=LOW,
        evidence=record.evidence,
        status_note=(
            "NEEDS REVIEW — contradiction: ultimate due $1,234.00 but "
            "reported PAID — flagged for human review"
        ),
        # All write_* deliberately None — the orchestrator's NEEDS_REVIEW
        # gate would not have populated them.
    )

    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, {f"s00_r{r:04d}": outcome}, RUN_DATE, out_path)
    ws = load_workbook(out_path)["Property Tax"]

    # Pay amt / date keep the template's pre-existing row-5 values byte
    # for byte. Actual assessment (AA) is analyst-owned and untouched.
    assert ws[f"Y{r}"].value == 19937
    assert ws[f"Z{r}"].value == "2025-11-15"
    assert ws[f"AA{r}"].value == 3666000
    # Confidence ALWAYS writes — even for NEEDS_REVIEW rows, since it's
    # informational and runs outside the `allowed` gate.
    assert ws[f"AB{r}"].value == "LOW"

    notes = ws.cell(r, _ptax_notes_col(out_path)).value
    assert "contradiction" in notes
    assert "flagged for human review" in notes


def test_low_confidence_writes_only_to_notes(ptax_master_workbook, tmp_path):
    # LOW confidence outcome — even with write_* set — must not touch
    # the structured cells (the `allowed` gate blocks every write).
    # The status note still reaches the notes column.
    r = PTAX_FIRST_DATA_ROW
    outcome = RowOutcome(
        row_key=f"s00_r{r:04d}",
        sheet_name="Property Tax",
        row_number=r,
        accounts=[AccountRecord(account_searched="x", status=PAID, confidence=LOW)],
        row_status=PAID,
        confidence=LOW,
        status_note="LOW confidence — could not verify",
        write_ultimate_payment_due=9999.99,  # legacy field — would be ignored anyway (header absent)
        write_payment_date="2099-12-31",     # gated by LOW confidence — must not write
        write_payment_amount=9999.99,        # ditto
        write_next_due_date="2099-12-31",    # legacy — header absent
    )

    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, {f"s00_r{r:04d}": outcome}, RUN_DATE, out_path)
    ws = load_workbook(out_path)["Property Tax"]

    assert ws[f"Y{r}"].value == 19937           # template's pre-existing value
    assert ws[f"Z{r}"].value == "2025-11-15"
    assert ws[f"AA{r}"].value == 3666000        # analyst column untouched
    assert ws[f"AB{r}"].value == "LOW"          # Confidence ALWAYS writes
    assert ws.cell(r, _ptax_notes_col(out_path)).value == (
        "LOW confidence — could not verify"
    )


def test_no_overwrite_with_blank(ptax_master_workbook, tmp_path):
    # Outcome with every write_* = None — a parse glitch that yielded no
    # values. The template's pre-existing row-5 values must survive
    # untouched (§7: no silent erasure).
    r = PTAX_FIRST_DATA_ROW
    outcome = RowOutcome(
        row_key=f"s00_r{r:04d}",
        sheet_name="Property Tax",
        row_number=r,
        accounts=[AccountRecord(account_searched="x", status=PAID, confidence=HIGH)],
        row_status=PAID,
        confidence=HIGH,
        status_note="Parse returned nothing — prior values preserved",
        # All write_* left None.
    )

    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, {f"s00_r{r:04d}": outcome}, RUN_DATE, out_path)
    ws = load_workbook(out_path)["Property Tax"]

    assert ws[f"Y{r}"].value == 19937
    assert ws[f"Z{r}"].value == "2025-11-15"
    assert ws[f"AA{r}"].value == 3666000        # analyst column untouched
    assert ws[f"AB{r}"].value == "HIGH"         # Confidence writes regardless


def test_date_correction_writes_and_notes(ptax_master_workbook, tmp_path):
    # Z5 (payment_date) is pre-populated as '2025-11-15' in the template.
    # A portal receipt of 2025-12-29 should (a) overwrite Z5 with a real
    # datetime, and (b) ride a "corrected payment date from ... to ...
    # per portal receipt" sentence into the notes column.
    r = PTAX_FIRST_DATA_ROW
    outcome = RowOutcome(
        row_key=f"s00_r{r:04d}",
        sheet_name="Property Tax",
        row_number=r,
        accounts=[AccountRecord(account_searched="x", status=PAID, confidence=HIGH)],
        row_status=PAID,
        confidence=HIGH,
        status_note="PAID in full $4,974.48 on 2025-12-29",
        write_payment_date="2025-12-29",
        write_payment_amount=4974.48,
    )

    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, {f"s00_r{r:04d}": outcome}, RUN_DATE, out_path)
    ws = load_workbook(out_path)["Property Tax"]

    assert ws[f"Z{r}"].value == dt.datetime(2025, 12, 29)
    assert ws[f"Y{r}"].value == 4974.48

    notes = ws.cell(r, _ptax_notes_col(out_path)).value
    assert "corrected payment date from 2025-11-15 to 2025-12-29 per portal receipt" in notes
    assert "corrected payment amount from 19937 to 4974.48 per portal receipt" in notes


def test_writeback_guard_blocks_protected_columns(ptax_master_workbook):
    # _assert_writable_column raises on any column whose detected
    # canonical field isn't in _WRITABLE_FIELDS (unless it's the notes
    # column). The error message names the offending header so a
    # misconfigured caller is obvious.
    intake = parse_workbook(ptax_master_workbook)
    sheet = intake.sheets[0]
    wb = load_workbook(ptax_master_workbook)
    ws = wb[sheet.name]

    # notes_col past every real column so its carve-out doesn't fire.
    notes_col = 999

    # Probe each analyst-owned column on the new 2026-06-25 layout:
    # BOV high (S=19), Jurisdiction link primary (V=22), Actual
    # assessment (AA=27 — fuzzy-matches assessed_value, which is NOT
    # in _WRITABLE_FIELDS so the guard still fires).
    from pontus_tax.intake import ColumnInfo
    for col in (19, 22, 27):
        # Build a ColumnInfo the way the writeback does — lift fieldname
        # off the parsed sheet so the test exercises the real path.
        info = None
        for cols in sheet.columns.values():
            for ci in cols:
                if ci.index == col:
                    info = ci
                    break
            if info:
                break
        # BOV high (S) and Jurisdiction primary (V) aren't mapped to any
        # canonical field at all — synthesize a bare ColumnInfo for them.
        if info is None:
            info = ColumnInfo(
                index=col, letter=get_column_letter(col),
                header=str(ws.cell(sheet.header_row, col).value),
                effective="", fieldname=None,
            )
        with pytest.raises(WritebackGuardError) as exc:
            _assert_writable_column(ws, sheet, info, notes_col)
        header = str(ws.cell(sheet.header_row, col).value)
        assert header in str(exc.value), (
            f"col {get_column_letter(col)}: error must name header "
            f"{header!r}, got: {exc.value!s}"
        )

    # Bare-int path: writing into the notes_col is always OK; writing
    # into any other column without a fieldname is not.
    _assert_writable_column(ws, sheet, notes_col, notes_col)
    with pytest.raises(WritebackGuardError):
        _assert_writable_column(ws, sheet, 27, notes_col)  # AA, not notes

    # Sanity: every writeback-owned canonical field on the current
    # template — Payment amount (Y), Payment date (Z), Confidence (AB) —
    # passes the guard.
    for fieldname, letter in (
        ("payment_amount", "Y"),
        ("payment_date",   "Z"),
        ("run_confidence", "AB"),
    ):
        info = sheet.first_col(fieldname)
        assert info is not None and info.letter == letter
        _assert_writable_column(ws, sheet, info, notes_col)


def test_template_analyst_columns_untouched(ptax_master_workbook, tmp_path):
    # Every analyst-owned column on the current PTAX_Master layout must
    # round-trip byte-for-byte through a real write_output run: every
    # (value, data_type) pair preserved. Catches accidental writes into
    # the BOV / Jurisdiction / Actual assessment columns.
    ws_before = load_workbook(ptax_master_workbook)["Property Tax"]
    # S–X = BOV high/mid/low + Jurisdiction primary/secondary/tertiary
    # AA = Actual assessment
    analyst_cols = list(range(19, 25)) + [27]
    before: dict[tuple[int, int], tuple] = {}
    for r in range(PTAX_FIRST_DATA_ROW, ws_before.max_row + 1):
        for col in analyst_cols:
            c = ws_before.cell(r, col)
            before[(r, col)] = (c.value, c.data_type)

    intake = parse_workbook(ptax_master_workbook)
    out_path = str(tmp_path / "out.xlsx")
    write_output(
        intake,
        {f"s00_r{PTAX_FIRST_DATA_ROW:04d}": _ptax_paid_putnam_outcome()},
        RUN_DATE, out_path,
    )

    ws_after = load_workbook(out_path)["Property Tax"]
    drifted: list[str] = []
    for (r, col), expected in before.items():
        c = ws_after.cell(r, col)
        if (c.value, c.data_type) != expected:
            drifted.append(
                f"{get_column_letter(col)}{r}: {expected} → "
                f"{(c.value, c.data_type)}"
            )
    assert not drifted, "analyst columns drifted:\n  " + "\n  ".join(drifted)


def test_unprocessed_rows_are_marked_not_skipped(florida_workbook, tmp_path):
    intake = parse_workbook(florida_workbook)
    outcomes = _outcomes()
    del outcomes["s00_r0005"]  # pretend the run died before row 5
    out_path = str(tmp_path / "out.xlsx")
    write_output(intake, outcomes, RUN_DATE, out_path)
    ws = load_workbook(out_path)["Florida Prop Tax"]
    # Notes column for the legacy Florida fixture is V (May 2026 Update).
    assert ws["V5"].value == "NOT CHECKED — run ended before this row"
