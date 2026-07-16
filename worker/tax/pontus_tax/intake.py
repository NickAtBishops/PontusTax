"""Workbook intake — CLAUDE.md §2.

Assume nothing about the workbook. Columns are detected by header text
(fuzzy/synonym matching, two stacked header rows supported), NEVER by
letter. Cells containing formulas are protected. Each sheet with a
detectable header row and at least one identifiable data row is processed.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .identifiers import AccountCandidates, split_accounts
from .taxonomy import domain_of

# IANA-reserved documentation/example domains (RFC 2606) plus localhost.
# These show up as leftover sample/placeholder data in reference columns
# an analyst didn't finish filling in (confirmed 2026-07-16: 29 of 80 rows
# in a real "with_urls" tracker had "https://example.com/x/property-N"
# style placeholders in Jurisdiction Link secondary/tertiary — visiting
# them produced correct-but-useless "not a property tax portal" results
# that dragged otherwise-good rows down to NEEDS_REVIEW). Never treat
# these as real portals to check — same as if the cell were empty.
_PLACEHOLDER_DOMAINS = {
    "example.com", "example.org", "example.net", "example.edu",
    "localhost", "test.com",
}


def _is_placeholder_domain(url: str) -> bool:
    domain = domain_of(url)
    if domain.startswith("www."):
        domain = domain[4:]
    return domain in _PLACEHOLDER_DOMAINS

# (canonical field, synonyms, word_match) — word_match synonyms only match as
# whole words (short tokens would otherwise fire inside unrelated headers).
SYNONYMS: list[tuple[str, list[str], bool]] = [
    ("internal_id", ["pid"], True),
    ("account_number", [
        "account number", "account #", "account", "acct", "parcel id",
        "parcel", "apn", "pin", "folio", "schedule", "tax id", "property id",
    ], False),
    ("address", ["property address", "address", "location", "situs"], False),
    ("city", ["city"], True),
    ("state", ["state", "st"], True),
    ("zip", ["zip code", "zip", "postal"], False),
    ("county", ["county", "parish", "borough", "taxing jurisdiction"], False),
    ("owner_entity", ["owner entity", "owner", "entity", "llc"], False),
    ("assessed_value", ["assessed value", "assessed", "assessment", "av"], True),
    ("tax_year", ["year of assessment", "tax year", "roll year", "year"], True),
    ("installments", ["# installments", "installments", "installment", "payments"], False),
    ("date_paid", ["date paid", "paid date", "paid on"], False),
    ("confirmation", ["paid confirmation", "confirmation", "receipt"], False),
    ("responsible_party", ["responsible", "tenant", "party"], False),
    ("website", ["website", "url", "link", "portal"], True),
    ("total", ["total"], True),
    ("status_notes", ["update", "notes", "status"], True),
    ("due_dates", ["due dates", "due date", "due"], True),
    ("amounts", ["amounts", "amount", "owed"], True),
    # "early bird" and "#1..#4" are typed by the data under them (§2.1):
    # date-typed → due_dates, number-typed → amounts. See _retype_numbered().
    ("_numbered", ["early bird", "#1", "#2", "#3", "#4", "#5", "#6"], False),
]

# Fields that may legitimately map to several columns.
MULTI_FIELDS = {"due_dates", "amounts", "status_notes"}

# PTAX_Master template (and future fixed-cell workbooks) names the
# structured cells the checker writes to. These headers are matched
# STRICTLY — case-insensitive and tolerant of extra whitespace, but NOT
# tolerant of paraphrase ("Amount due now" ≠ "Ultimate payment due").
# Strict matches always beat the fuzzy synonym table below, so e.g.
# "Payment amount" never falls through to the loose "amounts" bucket.
#
# The CURRENT PTAX_Master template (2026-06-25 reshuffle) only carries
# three of these: Payment amount, Payment date, Confidence — grouped under
# the "Model System Output" band. "Ultimate payment due" and "Next due
# date" were dropped from the layout; they stay in the table for backwards
# compatibility so older uploaded copies that still carry those headers
# continue to write into the right cell.
STRICT_HEADERS: dict[str, str] = {
    "payment amount":       "payment_amount",      # PTAX_Master Y
    "payment date":         "payment_date",        # PTAX_Master Z
    "confidence":           "run_confidence",      # PTAX_Master AB
    "ultimate payment due": "ultimate_payment_due",  # legacy template only
    "next due date":        "next_due_date",         # legacy template only
    # Jurisdiction Links band (PTAX_Master V/W/X). These headers contain the
    # word "link", which the fuzzy `website` synonym would otherwise match —
    # collapsing all three into the same bucket as "Website" and silently
    # dropping them (only the first-encountered column wins per fieldname
    # outside MULTI_FIELDS). Strict, exact-header matches give each of these
    # three its own canonical field so the scraping engine can visit them as
    # additional jurisdictions to check (analyst-filled; write-back never
    # writes to them — see writeback._WRITABLE_FIELDS).
    "jurisdiction link primary":   "jurisdiction_link_primary",
    "jurisdiction link secondary": "jurisdiction_link_secondary",
    "jurisdiction link tertiary":  "jurisdiction_link_tertiary",
}
_STRICT_SPECIFICITY = 1000  # any strict hit beats every fuzzy hit

UPDATE_PATTERN = re.compile(
    r"^(january|february|march|april|may|june|july|august|september|october|"
    r"november|december)\s+(\d{4})\s+update$",
    re.IGNORECASE,
)

URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)


@dataclass
class ColumnInfo:
    index: int            # 1-based
    letter: str
    header: str           # field-row text
    effective: str        # group + field text used for the match
    fieldname: str | None


@dataclass
class RowIntake:
    row_number: int
    address: str | None = None
    city: str | None = None
    state: str | None = None
    zip: str | None = None
    county: str | None = None
    owner_entity: str | None = None
    internal_id: str | None = None
    account_raw: str | None = None
    accounts: list[AccountCandidates] = field(default_factory=list)
    tax_year: str | None = None
    url: str | None = None
    # Jurisdiction Links band (PTAX_Master V/W/X) — analyst-filled reference
    # URLs for OTHER taxing authorities/portals this same property may owe
    # to (e.g. a county tax collector plus a special assessment district).
    # Read-only reference data: the write-back never writes to these (§10.5)
    # but the scraping engine checks each one that's filled in, in addition
    # to `url` (the "Website" column) — see `check_urls()` below.
    jurisdiction_link_primary: str | None = None
    jurisdiction_link_secondary: str | None = None
    jurisdiction_link_tertiary: str | None = None
    responsible_party: str | None = None
    date_paid_existing: Any = None
    confirmation_existing: Any = None
    assessed_existing: Any = None
    total_existing: Any = None
    amounts_existing: list[Any] = field(default_factory=list)

    @property
    def full_address(self) -> str:
        parts = [self.address, self.city, self.state, self.zip]
        return ", ".join(str(p).strip() for p in parts if p)

    def _url_candidates(self) -> list[tuple[str, str]]:
        return [
            ("Website", self.url),
            ("Jurisdiction link primary", self.jurisdiction_link_primary),
            ("Jurisdiction link secondary", self.jurisdiction_link_secondary),
            ("Jurisdiction link tertiary", self.jurisdiction_link_tertiary),
        ]

    def check_urls(self) -> list[tuple[str, str]]:
        """Every portal URL this row should be checked against, as ordered
        (label, url) pairs: "Website" first (when non-empty — an empty
        Website is a §4.6 discovery concern, not something this method
        fills in), then whichever of the three Jurisdiction Link columns
        carry a URL, in primary/secondary/tertiary order.

        A single property can owe tax to more than one taxing authority
        (e.g. county + a special assessment district), so every non-empty
        link is checked — not merely as a fallback when Website fails.

        Duplicate URLs (identical string, analyst pasted the same link into
        two columns) are checked only ONCE; the first-listed label wins.
        Placeholder/example URLs (see `_is_placeholder_domain`) are never
        checked at all — same as if the cell were empty; see
        `check_urls_skipped()` for the human-readable reason.
        """
        seen: set[str] = set()
        out: list[tuple[str, str]] = []
        for label, url in self._url_candidates():
            if not url:
                continue
            if _is_placeholder_domain(url):
                continue
            if url in seen:
                continue
            seen.add(url)
            out.append((label, url))
        return out

    def check_urls_skipped(self) -> list[str]:
        """Human-readable reasons for any URL this row carried that was
        NOT included in `check_urls()` because it looked like placeholder/
        example data — surfaced in the row's evidence so a bad reference
        link is visible to the analyst instead of silently vanishing."""
        notes: list[str] = []
        for label, url in self._url_candidates():
            if url and _is_placeholder_domain(url):
                notes.append(
                    f'{label} ("{url}") looks like placeholder/example data '
                    "(reserved example domain) — not a real portal, skipped"
                )
        return notes


@dataclass
class SheetIntake:
    name: str
    header_row: int
    group_row: int | None
    columns: dict[str, list[ColumnInfo]]
    rows: list[RowIntake]
    protected_columns: list[str]      # column letters that hold formulas
    ambiguous: list[str]
    update_columns: list[ColumnInfo]  # existing "<Month> <Year> Update" columns

    def first_col(self, fieldname: str) -> ColumnInfo | None:
        cols = self.columns.get(fieldname) or []
        return cols[0] if cols else None

    def mapping_doc(self) -> dict[str, str]:
        return {
            f: ",".join(c.letter for c in cols)
            for f, cols in sorted(self.columns.items())
            if cols
        }


@dataclass
class WorkbookIntake:
    path: str
    sheets: list[SheetIntake]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _match_field(text: str) -> tuple[str | None, int, list[str]]:
    """Best canonical field for a header text. Returns (field, specificity,
    tied_fields). STRICT_HEADERS map wins on exact (normalized) match;
    otherwise the longest fuzzy synonym wins. Ties across different fuzzy
    fields are reported as ambiguous."""
    t = " ".join(text.lower().split())
    if not t:
        return None, 0, []
    strict_hit = STRICT_HEADERS.get(t)
    if strict_hit is not None:
        return strict_hit, _STRICT_SPECIFICITY, []
    best: str | None = None
    best_len = 0
    tied: list[str] = []
    for fieldname, synonyms, word_match in SYNONYMS:
        for syn in synonyms:
            hit = (
                re.search(rf"(?<![a-z0-9]){re.escape(syn)}(?![a-z0-9])", t)
                if word_match or len(syn) <= 3
                else syn in t
            )
            if not hit:
                continue
            if len(syn) > best_len:
                best, best_len, tied = fieldname, len(syn), []
            elif len(syn) == best_len and fieldname != best and best is not None:
                tied.append(fieldname)
    return best, best_len, tied


def _group_text(ws, group_row: int | None, col: int) -> str:
    """Text of the group-header cell above a column, resolving merged ranges
    to their anchor cell."""
    if group_row is None:
        return ""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= group_row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return _clean(ws.cell(row=rng.min_row, column=rng.min_col).value)
    return _clean(ws.cell(row=group_row, column=col).value)


def _row_score(ws, row: int) -> int:
    fields = set()
    for col in range(1, min(ws.max_column, 60) + 1):
        f, _, _ = _match_field(_clean(ws.cell(row=row, column=col).value))
        if f and f != "_numbered":
            fields.add(f)
    return len(fields)


def detect_header(ws) -> tuple[int | None, int | None]:
    """Scan the first ~5 rows for the field-header row; the row above is the
    group row when it carries text but matches fewer fields (Florida has a
    group row stacked above the real header row)."""
    scores = {r: _row_score(ws, r) for r in range(1, min(6, ws.max_row + 1))}
    if not scores:
        return None, None
    header_row = max(scores, key=lambda r: (scores[r], -r))
    if scores[header_row] < 3:
        return None, None
    group_row = None
    if header_row > 1:
        above = header_row - 1
        has_text = any(
            _clean(ws.cell(row=above, column=c).value)
            for c in range(1, min(ws.max_column, 60) + 1)
        )
        if has_text and scores.get(above, 0) < scores[header_row]:
            group_row = above
    return header_row, group_row


def _sample_type(ws_values, col: int, start: int, end: int) -> str:
    """Type the data under a numbered/'early bird' header: dates → due_dates,
    numbers → amounts."""
    dates = numbers = 0
    for r in range(start, min(end, start + 30) + 1):
        v = ws_values.cell(row=r, column=col).value
        if isinstance(v, (dt.datetime, dt.date)):
            dates += 1
        elif isinstance(v, (int, float)):
            numbers += 1
        elif isinstance(v, str):
            s = v.strip()
            if re.fullmatch(r"\$?[\d,]+(\.\d+)?", s):
                numbers += 1
            elif re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", s):
                dates += 1
    if dates == 0 and numbers == 0:
        return "amounts"
    return "due_dates" if dates >= numbers else "amounts"


def _extract_url(cell) -> str | None:
    if cell.hyperlink is not None and cell.hyperlink.target:
        return str(cell.hyperlink.target)
    text = _clean(cell.value)
    if not text:
        return None
    m = re.search(r'=HYPERLINK\(\s*"([^"]+)"', text, re.IGNORECASE)
    if m:
        return m.group(1)
    m = URL_RE.search(text)
    if m:
        return m.group(0).rstrip(").,;")
    if re.match(r"^(www\.)?[\w.-]+\.(gov|us|com|org|net)(/\S*)?$", text):
        return f"https://{text}" if not text.startswith("www.") else f"https://{text}"
    return None


def parse_workbook(path: str) -> WorkbookIntake:
    wb_f = load_workbook(path, data_only=False)   # formulas intact
    wb_v = load_workbook(path, data_only=True)    # cached values
    sheets: list[SheetIntake] = []

    for ws in wb_f.worksheets:
        if ws.sheet_state != "visible":
            continue
        ws_v = wb_v[ws.title]
        header_row, group_row = detect_header(ws)
        if header_row is None:
            continue

        ambiguous: list[str] = []
        infos: list[ColumnInfo] = []
        for col in range(1, ws.max_column + 1):
            header = _clean(ws.cell(row=header_row, column=col).value)
            group = _group_text(ws, group_row, col)
            fieldname, spec, tied = _match_field(header)
            effective = header
            if fieldname is None:
                combined = f"{group} {header}".strip()
                if combined and combined.lower() != header.lower():
                    fieldname, spec, tied = _match_field(combined)
                    effective = combined
            letter = get_column_letter(col)
            if tied:
                ambiguous.append(
                    f"column {letter} ('{effective}') matched several fields: "
                    f"{fieldname}, {', '.join(tied)} — using {fieldname}"
                )
            if not header and not group:
                fieldname = None
            infos.append(ColumnInfo(col, letter, header, effective, fieldname))

        data_start = header_row + 1
        data_end = ws.max_row

        # Resolve "#N"/"early bird" columns and anything that landed on the
        # generic due/amount buckets by the type of the data beneath (§2.1).
        for info in infos:
            if info.fieldname == "_numbered":
                info.fieldname = _sample_type(ws_v, info.index, data_start, data_end)

        columns: dict[str, list[ColumnInfo]] = {}
        for info in infos:
            if not info.fieldname:
                continue
            bucket = columns.setdefault(info.fieldname, [])
            if info.fieldname not in MULTI_FIELDS and bucket:
                ambiguous.append(
                    f"column {info.letter} ('{info.effective}') also matched "
                    f"{info.fieldname}; keeping {bucket[0].letter}"
                )
                continue
            bucket.append(info)

        # Formula protection (§2.1): a column whose data cells are MOSTLY
        # formulas is protected as a whole (Florida's Total column W). A
        # single SUM in a totals row must NOT freeze an otherwise writable
        # column — per-cell guards in write-back handle those.
        protected: set[str] = set()
        for info in infos:
            formulas = nonempty = 0
            for r in range(data_start, data_end + 1):
                c = ws.cell(row=r, column=info.index)
                if c.value is None:
                    continue
                nonempty += 1
                if c.data_type == "f" or (
                    isinstance(c.value, str) and c.value.startswith("=")
                ):
                    formulas += 1
            if formulas and formulas >= nonempty * 0.5:
                protected.add(info.letter)

        def val(fieldname: str, row: int) -> Any:
            info = columns.get(fieldname, [None])[0] if columns.get(fieldname) else None
            if info is None:
                return None
            return ws_v.cell(row=row, column=info.index).value

        rows: list[RowIntake] = []
        for r in range(data_start, data_end + 1):
            account_raw = _clean(val("account_number", r)) or None
            url = None
            web_info = columns.get("website", [None])[0] if columns.get("website") else None
            if web_info is not None:
                url = _extract_url(ws.cell(row=r, column=web_info.index))
            jl_primary = jl_secondary = jl_tertiary = None
            jl_primary_info = (
                columns.get("jurisdiction_link_primary", [None])[0]
                if columns.get("jurisdiction_link_primary") else None
            )
            if jl_primary_info is not None:
                jl_primary = _extract_url(ws.cell(row=r, column=jl_primary_info.index))
            jl_secondary_info = (
                columns.get("jurisdiction_link_secondary", [None])[0]
                if columns.get("jurisdiction_link_secondary") else None
            )
            if jl_secondary_info is not None:
                jl_secondary = _extract_url(ws.cell(row=r, column=jl_secondary_info.index))
            jl_tertiary_info = (
                columns.get("jurisdiction_link_tertiary", [None])[0]
                if columns.get("jurisdiction_link_tertiary") else None
            )
            if jl_tertiary_info is not None:
                jl_tertiary = _extract_url(ws.cell(row=r, column=jl_tertiary_info.index))
            address = _clean(val("address", r)) or None
            if not (account_raw or url or address):
                continue  # totals row, spacer, narrative row
            year_val = val("tax_year", r)
            amounts_existing = [
                ws_v.cell(row=r, column=c.index).value
                for c in columns.get("amounts", [])
            ]
            rows.append(
                RowIntake(
                    row_number=r,
                    address=address,
                    city=_clean(val("city", r)) or None,
                    state=_clean(val("state", r)) or None,
                    zip=_clean(val("zip", r)) or None,
                    county=_clean(val("county", r)) or None,
                    owner_entity=_clean(val("owner_entity", r)) or None,
                    internal_id=_clean(val("internal_id", r)) or None,
                    account_raw=account_raw,
                    accounts=split_accounts(account_raw),
                    tax_year=_clean(year_val) or None,
                    url=url,
                    jurisdiction_link_primary=jl_primary,
                    jurisdiction_link_secondary=jl_secondary,
                    jurisdiction_link_tertiary=jl_tertiary,
                    responsible_party=_clean(val("responsible_party", r)) or None,
                    date_paid_existing=val("date_paid", r),
                    confirmation_existing=val("confirmation", r),
                    assessed_existing=val("assessed_value", r),
                    total_existing=val("total", r),
                    amounts_existing=amounts_existing,
                )
            )

        if not rows:
            continue
        if not any(row.url or row.accounts for row in rows):
            continue  # §2.1: need at least one data row with a URL or account

        update_columns = [
            c for c in columns.get("status_notes", [])
            if UPDATE_PATTERN.match(c.header.strip())
        ]
        sheets.append(
            SheetIntake(
                name=ws.title,
                header_row=header_row,
                group_row=group_row,
                columns=columns,
                rows=rows,
                protected_columns=sorted(protected),
                ambiguous=ambiguous,
                update_columns=update_columns,
            )
        )

    return WorkbookIntake(path=path, sheets=sheets)


NOTES_HEADER_DEFAULT = "Last run notes"


def status_column_header(intake_sheet: SheetIntake) -> str:
    """Header text of the notes column write-back will overwrite this run.

    The new fixed-layout model (§10) NEVER appends a per-run column.
    Instead it overwrites a single notes cell:
      - The rightmost existing '<Month> YYYY Update' column if any
        (older Florida-style workbooks); the historical month-stamped
        columns to its left are left untouched as artifacts.
      - Otherwise a single 'Last run notes' column the writeback
        creates exactly once on first run.
    """
    if intake_sheet.update_columns:
        return intake_sheet.update_columns[-1].header
    return NOTES_HEADER_DEFAULT
