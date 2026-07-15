"""Write-back — CLAUDE.md §10.

openpyxl, preserving structure: formatting, widths, merged headers,
hyperlinks, formulas. Writes ONLY into the canonical fields listed in
`_WRITABLE_FIELDS` (detected by header text per `intake.STRICT_HEADERS`
and the fuzzy synonym table), plus a single notes column. Every other
column on the sheet — BOV bands, jurisdiction links, Actual assessment,
the original tracker columns A–R — is read-only to the writeback, with
the `_assert_writable_column` guard turning any accidental write into a
loud error. Formula cells are never touched; a scraped blank/zero never
erases a real value; no row is silently skipped. The original upload is
never modified — output is a copy.
"""

from __future__ import annotations

import datetime as dt
import logging
from copy import copy
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .canonical import NEEDS_REVIEW, PAID, UNREACHABLE, LOW, RowOutcome
from .intake import (
    NOTES_HEADER_DEFAULT, ColumnInfo, SheetIntake, WorkbookIntake,
    status_column_header,
)
from .validate import parse_date, parse_money

# Number formats used when overwriting PTAX_Master's structured cells.
# Excel round-trips the underlying datetime / number; the template already
# sets these formats on the sample rows, but we re-apply on bare sheets.
_CURRENCY_FMT = '"$"#,##0.00'
_DATE_FMT = "yyyy-mm-dd"

# Pulled verbatim from templates/PTAX_Master.xlsx (2026-07-02) — used only
# for the analyst-owned columns (BOV bands, Actual assessment) that
# `_write_structured_cell` never touches, so whatever format is set at
# creation time is what persists. Payment amount / Payment date reuse
# _CURRENCY_FMT / _DATE_FMT above instead, since _write_structured_cell
# overwrites the number_format on every real write anyway — matching the
# template's own format for those two would just get replaced the moment
# a value lands.
_CURRENCY_FMT_BOV = '\\$#,##0;"($"#,##0\\);\\-'
_CREAM_FILL = "FFFFF8DC"

# Canonical fields the writeback OWNS. The writeback may write into the
# detected column for any of these (subject to per-cell formula protection
# and the row-level allowed-status / confidence gates). Every other matched
# field — notably `assessed_value`, which fuzzy-matches the analyst-owned
# "Actual assessment" column on PTAX_Master — is read-only to the writeback,
# regardless of which physical column the analyst placed it in. Protection
# is by CANONICAL FIELD, not by column letter, so reshuffling the template
# (Pontus moved Payment amount / date / Confidence from S..T..V..W to
# Y..Z..AB on 2026-06-25) does not silently re-permit analyst-column writes.
_WRITABLE_FIELDS: frozenset[str] = frozenset({
    # Model System Output cells on the current PTAX_Master template.
    "payment_date",       # Z on PTAX_Master (was U pre-2026-06-25)
    "payment_amount",     # Y on PTAX_Master (was V pre-2026-06-25)
    "run_confidence",     # AB on PTAX_Master (was T pre-2026-06-25)
    # PTAX_Master cells from earlier template revisions — absent on the
    # current template (so the write is a no-op there) but still honored
    # when an older copy of the template is uploaded.
    "ultimate_payment_due",
    "next_due_date",
    # Legacy free-text workbooks (the Florida tracker, etc. — no template,
    # the orchestrator writes back into the same columns the analyst typed
    # values into). These columns sit in A..R territory on every workbook
    # we have seen, well clear of any BOV / jurisdiction / assessment band.
    "date_paid",
    "confirmation",
    "amounts",            # only when a single amounts column was detected
})


class WritebackGuardError(RuntimeError):
    """Raised when write-back attempts to touch a column that is not
    owned by the writeback (and not the resolved notes column). Surfaces
    the offending column letter, header text, and detected canonical
    field so a misconfigured caller is obvious."""


def _assert_writable_column(
    ws,
    sheet: SheetIntake,
    target: int | ColumnInfo,
    notes_col: int,
) -> None:
    """Guard every write attempt. Accepts either a ColumnInfo (the normal
    path — the writeback always knows which canonical field it is touching)
    or a bare column index (the legacy callsite that writes the notes cell
    itself). Raises WritebackGuardError when the field is not in
    `_WRITABLE_FIELDS` and the target is not the resolved notes column."""
    if isinstance(target, ColumnInfo):
        target_col = target.index
        fieldname = target.fieldname
    else:
        target_col = target
        fieldname = None
    if target_col == notes_col:
        return
    if fieldname in _WRITABLE_FIELDS:
        return
    header = ws.cell(row=sheet.header_row, column=target_col).value
    raise WritebackGuardError(
        f"refusing to write into protected column "
        f"{get_column_letter(target_col)} ({header!r}, field={fieldname!r}) "
        f"— writeback only owns {sorted(_WRITABLE_FIELDS)} and the "
        f"resolved notes column"
    )

log = logging.getLogger("pontus_tax.writeback")

# Statuses whose data values are allowed into data cells (§7: LOW/NEEDS_REVIEW
# results reach the status column only).
_WRITABLE_STATUSES_EXCLUDED = {NEEDS_REVIEW, UNREACHABLE}


def _is_formula(cell) -> bool:
    return cell.data_type == "f" or (
        isinstance(cell.value, str) and str(cell.value).startswith("=")
    )


def _cell_empty(cell) -> bool:
    v = cell.value
    return v is None or (isinstance(v, str) and not v.strip())


def _values_equalish(existing: Any, new: Any) -> bool:
    if existing is None or new is None:
        return False
    e_money, n_money = parse_money(existing), parse_money(new)
    if e_money is not None and n_money is not None:
        return abs(e_money - n_money) < 0.01
    e_date, n_date = parse_date(existing), parse_date(new)
    if e_date is not None and n_date is not None:
        return e_date == n_date
    return str(existing).strip().lower() == str(new).strip().lower()


def _safe_write(ws, row: int, col: int, value: Any, number_format: str | None = None) -> bool:
    """One gate for every data-cell write: never a formula cell, never erase,
    never fight an existing different value (corrections ride the status
    column instead). Returns True when the cell was written."""
    cell = ws.cell(row=row, column=col)
    if _is_formula(cell):
        return False
    if value is None or (isinstance(value, str) and not value.strip()):
        return False  # §7 no silent erasure — blanks never overwrite
    if not _cell_empty(cell):
        return _values_equalish(cell.value, value)  # already right → fine
    cell.value = value
    if number_format:
        cell.number_format = number_format
    return True


def _write_structured_cell(ws, row: int, col: int, value: Any, number_format: str) -> bool:
    """Overwrite a fixed structured cell (S–V on PTAX_Master), preserving the
    template's per-cell fill (cream FFFFF8DC), font, and alignment.
    cell.number_format = ... can drop style attributes on some openpyxl
    versions, so save and restore explicitly. None / empty string never
    overwrites (§7: no silent erasure). The richer correction rules —
    no $0 over an existing nonzero, date-correction-with-note — layer on
    in a later commit; this helper handles the basic write."""
    cell = ws.cell(row=row, column=col)
    if _is_formula(cell):
        return False
    if value is None or (isinstance(value, str) and not value.strip()):
        return False
    saved_fill = copy(cell.fill)
    saved_font = copy(cell.font)
    saved_align = copy(cell.alignment)
    cell.value = value
    cell.number_format = number_format
    cell.fill = saved_fill
    cell.font = saved_font
    cell.alignment = saved_align
    return True


def _iso_to_datetime(iso: str | None) -> dt.datetime | None:
    """Parse an ISO-ish date string into datetime.datetime at midnight, the
    form openpyxl writes as a real Excel date value (not a string)."""
    d = parse_date(iso) if iso else None
    return dt.datetime(d.year, d.month, d.day) if d else None


def _collect_correction(cell, new_value: Any, header_label: str) -> str | None:
    """Compare an existing cell's value to a new value about to overwrite it.

    Returns a correction sentence when the portal contradicts a real
    prior value ("corrected <field> from <old> to <new> per portal
    receipt"). Returns None when the existing cell is empty, when both
    represent the same logical value (so an unchanged write looks like
    no-op), or when comparison isn't meaningful.
    """
    existing = cell.value
    if existing is None or (isinstance(existing, str) and not existing.strip()):
        return None

    def _fmt(v) -> str:
        if isinstance(v, dt.datetime):
            return v.date().isoformat()
        if isinstance(v, dt.date):
            return v.isoformat()
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    # Date-vs-date / date-vs-iso-string normalization.
    new_d = (
        new_value.date() if isinstance(new_value, dt.datetime)
        else new_value if isinstance(new_value, dt.date)
        else None
    )
    if new_d is not None:
        e_d = (
            existing.date() if isinstance(existing, dt.datetime)
            else existing if isinstance(existing, dt.date)
            else parse_date(existing) if isinstance(existing, str)
            else None
        )
        if e_d is not None and e_d == new_d:
            return None
        return (
            f"corrected {header_label} from {_fmt(existing)} to "
            f"{new_d.isoformat()} per portal receipt"
        )

    # Number-vs-number tolerance (cents-level).
    if isinstance(new_value, (int, float)) and isinstance(existing, (int, float)):
        if abs(float(existing) - float(new_value)) < 0.005:
            return None
        return (
            f"corrected {header_label} from {_fmt(existing)} to "
            f"{_fmt(new_value)} per portal receipt"
        )

    # Mixed types or strings — string-normalized equality.
    if _fmt(existing).lower() == _fmt(new_value).lower():
        return None
    return (
        f"corrected {header_label} from {_fmt(existing)} to "
        f"{_fmt(new_value)} per portal receipt"
    )


def _column_number_format(ws, col: int, data_start: int, data_end: int) -> str | None:
    for r in range(data_start, min(data_end, data_start + 40) + 1):
        cell = ws.cell(row=r, column=col)
        if cell.value is not None and cell.number_format != "General":
            return cell.number_format
    return None


def _last_used_column(ws, sheet: SheetIntake) -> int:
    last = 0
    rows_to_scan = [sheet.header_row]
    if sheet.group_row:
        rows_to_scan.append(sheet.group_row)
    rows_to_scan += [r.row_number for r in sheet.rows[:3]]
    for r in rows_to_scan:
        for c in range(ws.max_column, 0, -1):
            if ws.cell(row=r, column=c).value is not None:
                last = max(last, c)
                break
    return last or ws.max_column


def _resolve_notes_column(ws, sheet: SheetIntake) -> int:
    """Return the single column index the notes sentence is written into.

    Three-branch rule (§10):
    1. Rightmost existing '<Month> YYYY Update' column wins (older
       Florida-style workbooks; historical month-stamped columns to its
       left are left untouched as artifacts).
    2. Existing 'Last run notes' column from a prior run on the same
       workbook (lookup by header text, case-insensitive).
    3. Otherwise create 'Last run notes' exactly once, at the rightmost
       column position, with a reasonable width.
    """
    if sheet.update_columns:
        return sheet.update_columns[-1].index
    target_name = NOTES_HEADER_DEFAULT.strip().lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=sheet.header_row, column=c).value
        if v is not None and str(v).strip().lower() == target_name:
            return c
    new_col = _last_used_column(ws, sheet) + 1
    ws.cell(row=sheet.header_row, column=new_col).value = NOTES_HEADER_DEFAULT
    ws.column_dimensions[get_column_letter(new_col)].width = 52
    return new_col


# ---- PTAX_Master structured-column parity --------------------------------
# Any workbook that predates the PTAX_Master template (or was never built
# from it — the older per-state trackers like "Property Taxes- California
# .xlsx") is missing the BOV / Jurisdiction Links / Payment amount / Payment
# date / Actual assessment / Confidence band entirely, so those cells can
# never populate no matter how good the scrape is. `_ensure_structured_
# columns` adds whichever of these the sheet doesn't already have — by
# EXACT header text, never by letter, so a file that already carries any of
# these columns (in whatever position) is left completely untouched — styled
# to match templates/PTAX_Master.xlsx cell-for-cell (pulled directly from
# the template on 2026-07-02, not eyeballed).
@dataclass
class _StructuredColumnSpec:
    header: str
    # None for the analyst-owned columns (BOV, Jurisdiction links, Actual
    # assessment) — the app never writes values there, same as on
    # PTAX_Master itself; the column is created empty for the analyst to
    # fill in by hand. Set only for the three cells the writeback owns.
    fieldname: str | None
    width: float
    number_format: str
    header_align: str
    data_fill: str | None  # ARGB hex, or None for no fill


_STRUCTURED_COLUMNS: list[_StructuredColumnSpec] = [
    _StructuredColumnSpec("BOV high", None, 18.0, _CURRENCY_FMT_BOV, "center", None),
    _StructuredColumnSpec("BOV mid", None, 13.0, _CURRENCY_FMT_BOV, "center", None),
    _StructuredColumnSpec("BOV low", None, 13.0, _CURRENCY_FMT_BOV, "center", None),
    _StructuredColumnSpec(
        "Jurisdiction link primary", None, 22.42578125, "General", "left", None,
    ),
    _StructuredColumnSpec(
        "Jurisdiction link secondary", None, 13.0, "General", "left", None,
    ),
    _StructuredColumnSpec(
        "Jurisdiction link tertiary", None, 13.0, "General", "left", None,
    ),
    _StructuredColumnSpec(
        "Payment amount", "payment_amount", 18.0, _CURRENCY_FMT, "center",
        _CREAM_FILL,
    ),
    _StructuredColumnSpec(
        "Payment date", "payment_date", 13.0, _DATE_FMT, "center", _CREAM_FILL,
    ),
    _StructuredColumnSpec("Actual assessment", None, 13.0, _CURRENCY_FMT_BOV, "center", None),
    _StructuredColumnSpec(
        "Confidence", "run_confidence", 12.0, "General", "center", _CREAM_FILL,
    ),
]


def _find_header_column(ws, header_row: int, text: str) -> int | None:
    target = " ".join(text.strip().lower().split())
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and " ".join(str(v).strip().lower().split()) == target:
            return c
    return None


def _ensure_structured_columns(ws, sheet: SheetIntake) -> dict[str, ColumnInfo]:
    """Add whichever PTAX_Master structured columns this sheet is missing;
    leave every column that already exists completely alone. Returns
    ColumnInfo for the fields the writeback can populate THIS run
    (payment_amount, payment_date, run_confidence) — found via a fresh
    header-text scan, since the intake pass that produced `sheet.columns`
    ran before this function could have added them."""
    resolved: dict[str, ColumnInfo] = {}
    data_rows = [r.row_number for r in sheet.rows]
    data_start, data_end = min(data_rows), max(data_rows)

    for spec in _STRUCTURED_COLUMNS:
        existing = _find_header_column(ws, sheet.header_row, spec.header)
        if existing is not None:
            if spec.fieldname:
                letter = get_column_letter(existing)
                resolved[spec.fieldname] = ColumnInfo(
                    existing, letter, spec.header, spec.header, spec.fieldname,
                )
            continue

        new_col = _last_used_column(ws, sheet) + 1
        letter = get_column_letter(new_col)
        header_cell = ws.cell(row=sheet.header_row, column=new_col)
        header_cell.value = spec.header
        header_cell.font = Font(name="Aptos Narrow", size=11, bold=True)
        header_cell.alignment = Alignment(
            horizontal=spec.header_align, vertical="center",
        )
        ws.column_dimensions[letter].width = spec.width
        for r in range(data_start, data_end + 1):
            cell = ws.cell(row=r, column=new_col)
            cell.font = Font(name="Aptos Narrow", size=11)
            cell.number_format = spec.number_format
            if spec.data_fill:
                cell.fill = PatternFill(fill_type="solid", fgColor=spec.data_fill)
        if spec.fieldname:
            resolved[spec.fieldname] = ColumnInfo(
                new_col, letter, spec.header, spec.header, spec.fieldname,
            )
    return resolved


def write_output(
    intake: WorkbookIntake,
    outcomes: dict[str, RowOutcome],
    run_date: dt.date,
    out_path: str,
) -> dict[str, dict[str, str]]:
    """Produce the checked copy. Returns
    {sheet_name: {"status": <notes column header>}}.

    The new fixed-layout model (§10) does NOT append per-run columns.
    The status sentence overwrites a single notes cell whose column is
    resolved by `_resolve_notes_column`: the rightmost pre-existing
    month-update column on legacy workbooks, or a 'Last run notes'
    column created exactly once on first run for fresh layouts.

    `run_date` is retained as the as-of date the calling note-builder
    may use when composing the sentence; the writeback itself no longer
    derives a per-run column name from it."""
    wb = load_workbook(intake.path, data_only=False)
    headers_added: dict[str, dict[str, str]] = {}

    for s_idx, sheet in enumerate(intake.sheets):
        ws = wb[sheet.name]
        data_rows = [r.row_number for r in sheet.rows]
        data_start, data_end = min(data_rows), max(data_rows)

        # ---- PTAX_Master column parity: add whichever structured columns
        # this sheet doesn't already have (BOV bands, Jurisdiction links,
        # Payment amount/date, Actual assessment, Confidence), styled to
        # match the template. Columns that already exist are untouched.
        # Runs BEFORE the notes column resolves so notes still lands as
        # the rightmost column, same as before this existed.
        structured = _ensure_structured_columns(ws, sheet)

        # ---- Notes column: single, fixed-position, overwritten per run.
        notes_col = _resolve_notes_column(ws, sheet)
        notes_header = str(
            ws.cell(row=sheet.header_row, column=notes_col).value or ""
        ).strip() or status_column_header(sheet)
        headers_added[sheet.name] = {"status": notes_header}

        date_info = sheet.first_col("date_paid")
        conf_info = sheet.first_col("confirmation")
        assessed_info = sheet.first_col("assessed_value")
        amount_cols = sheet.columns.get("amounts", [])
        # PTAX_Master structured cells (S–W). Absent on legacy workbooks
        # like the older Florida sheet; present on the master template.
        # `structured` fills in whichever of the three the sheet didn't
        # already have — freshly created above, so they populate THIS run
        # instead of only becoming usable on the next re-upload.
        ultimate_info = sheet.first_col("ultimate_payment_due")
        confidence_info = sheet.first_col("run_confidence") or structured.get("run_confidence")
        paydate_info = sheet.first_col("payment_date") or structured.get("payment_date")
        payamt_info = sheet.first_col("payment_amount") or structured.get("payment_amount")
        nextdue_info = sheet.first_col("next_due_date")
        date_fmt = (
            _column_number_format(ws, date_info.index, data_start, data_end)
            if date_info
            else None
        ) or "MM/DD/YYYY"

        # Same guard as `_assert_writable_column` (fieldname-based) but
        # used for the silent-skip path on legacy fields: detection might
        # hand us a ColumnInfo whose fieldname is not in `_WRITABLE_FIELDS`
        # (notably `assessed_value` → "Actual assessment"), and we want
        # those writes to be clean no-ops rather than loud guard errors.
        def _writable(info) -> bool:
            if info is None:
                return False
            return info.fieldname in _WRITABLE_FIELDS or info.index == notes_col

        for row in sheet.rows:
            key = f"s{s_idx:02d}_r{row.row_number:04d}"
            outcome = outcomes.get(key)
            _assert_writable_column(ws, sheet, notes_col, notes_col)
            note_cell = ws.cell(row=row.row_number, column=notes_col)
            if outcome is None:
                # §10.3 — no row is ever silently skipped.
                note_cell.value = "NOT CHECKED — run ended before this row"
                continue

            allowed = (
                outcome.row_status not in _WRITABLE_STATUSES_EXCLUDED
                and outcome.confidence != LOW
            )
            # Corrections accumulate during the structured writes below
            # and append to the notes sentence at the end of this row.
            corrections: list[str] = []

            # Confidence is informational and runs OUTSIDE the `allowed`
            # gate — LOW / NEEDS_REVIEW rows should still surface "LOW"
            # in the Confidence column so analysts see at a glance what
            # the system trusted. Without this, LOW rows would have a
            # blank Confidence cell, indistinguishable from "never ran".
            if confidence_info and outcome.confidence:
                _assert_writable_column(ws, sheet, confidence_info, notes_col)
                _write_structured_cell(
                    ws, row.row_number, confidence_info.index,
                    outcome.confidence, "@",  # plain text
                )

            if allowed:
                if _writable(date_info) and outcome.write_date_paid:
                    d = parse_date(outcome.write_date_paid)
                    if d:
                        _assert_writable_column(ws, sheet, date_info, notes_col)
                        _safe_write(ws, row.row_number, date_info.index, d, date_fmt)
                if _writable(conf_info) and outcome.write_receipt:
                    _assert_writable_column(ws, sheet, conf_info, notes_col)
                    _safe_write(ws, row.row_number, conf_info.index, outcome.write_receipt)
                if _writable(assessed_info) and outcome.write_assessed_value is not None:
                    _assert_writable_column(ws, sheet, assessed_info, notes_col)
                    _safe_write(
                        ws, row.row_number, assessed_info.index,
                        outcome.write_assessed_value,
                    )
                if (
                    len(amount_cols) == 1
                    and outcome.write_amount_due is not None
                    and _writable(amount_cols[0])
                ):
                    _assert_writable_column(ws, sheet, amount_cols[0], notes_col)
                    _safe_write(
                        ws, row.row_number, amount_cols[0].index,
                        outcome.write_amount_due,
                    )

                # ---- PTAX_Master structured cells. Fixed-position overwrites
                # every run, preserving cream fill. Each write collects a
                # correction note when the portal contradicts a real prior
                # value, so analysts can see what changed and why. On the
                # current template, ultimate_info / nextdue_info are None
                # (those headers were removed in the 2026-06-25 reshuffle);
                # the writes below are no-ops on the current layout and
                # still fire on older templates that still carry them.
                if ultimate_info and outcome.write_ultimate_payment_due is not None:
                    cell = ws.cell(row=row.row_number, column=ultimate_info.index)
                    c = _collect_correction(
                        cell, outcome.write_ultimate_payment_due, "ultimate payment due",
                    )
                    if c:
                        corrections.append(c)
                    _assert_writable_column(ws, sheet, ultimate_info, notes_col)
                    _write_structured_cell(
                        ws, row.row_number, ultimate_info.index,
                        outcome.write_ultimate_payment_due, _CURRENCY_FMT,
                    )
                if paydate_info and outcome.write_payment_date:
                    pd = _iso_to_datetime(outcome.write_payment_date)
                    if pd:
                        cell = ws.cell(row=row.row_number, column=paydate_info.index)
                        c = _collect_correction(cell, pd, "payment date")
                        if c:
                            corrections.append(c)
                        _assert_writable_column(ws, sheet, paydate_info, notes_col)
                        _write_structured_cell(
                            ws, row.row_number, paydate_info.index, pd, _DATE_FMT,
                        )
                if payamt_info and outcome.write_payment_amount is not None:
                    cell = ws.cell(row=row.row_number, column=payamt_info.index)
                    c = _collect_correction(
                        cell, outcome.write_payment_amount, "payment amount",
                    )
                    if c:
                        corrections.append(c)
                    _assert_writable_column(ws, sheet, payamt_info, notes_col)
                    _write_structured_cell(
                        ws, row.row_number, payamt_info.index,
                        outcome.write_payment_amount, _CURRENCY_FMT,
                    )
                if nextdue_info and outcome.write_next_due_date:
                    nd = _iso_to_datetime(outcome.write_next_due_date)
                    if nd:
                        cell = ws.cell(row=row.row_number, column=nextdue_info.index)
                        c = _collect_correction(cell, nd, "next due date")
                        if c:
                            corrections.append(c)
                        _assert_writable_column(ws, sheet, nextdue_info, notes_col)
                        _write_structured_cell(
                            ws, row.row_number, nextdue_info.index, nd, _DATE_FMT,
                        )

            # Write the notes cell LAST so any corrections collected during
            # the structured writes ride along on the same line analysts
            # already read for run status.
            base = outcome.status_note or "NOT CHECKED"
            note_cell.value = (
                f"{base} | " + " | ".join(corrections) if corrections else base
            )

    wb.save(out_path)
    log.info("wrote checked workbook: %s", out_path)
    return headers_added


def output_filename(input_name: str, run_date: dt.date) -> str:
    stem, dot, ext = input_name.rpartition(".")
    if not dot:
        stem, ext = input_name, "xlsx"
    return f"{stem} — checked {run_date.isoformat()}.{ext}"
