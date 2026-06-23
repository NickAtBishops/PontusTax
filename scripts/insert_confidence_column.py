"""One-shot: insert a "Confidence" column at T in PTAX_Master.xlsx.

Run once: rewrites templates/PTAX_Master.xlsx in place. The new column
sits immediately right of "Ultimate payment due" (S) and pushes every
analyst-filled / formula column W..AJ one to the right (X..AK after).

Three things openpyxl's insert_cols does NOT do for you:
  1) Rewrite formula text that references shifted columns by letter
     (variance chain in AH/AI/AJ → AI/AJ/AK).
  2) Shift row-2 group-header merged ranges past the insert point.
  3) Extend a straddling merge (the Payment Status band, S2:V2, must
     grow to S2:W2 to absorb the new T column).

This script handles all three.
"""
from __future__ import annotations

import os
import re
import shutil
import sys
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

TEMPLATE = os.path.join(
    os.path.dirname(__file__), "..", "templates", "PTAX_Master.xlsx"
)
TEMPLATE = os.path.abspath(TEMPLATE)
BACKUP = TEMPLATE + ".pre-confidence.bak"

INSERT_AT = column_index_from_string("T")  # 20 — between S and the old T
CREAM = "FFFFF8DC"
CONFIDENCE_HEADER = "Confidence"


def _shift_letter(letter: str, insert_at: int) -> str:
    idx = column_index_from_string(letter)
    return get_column_letter(idx + 1 if idx >= insert_at else idx)


def _rewrite_formula(formula: str, insert_at: int) -> str:
    """Bump every column-letter reference that's at or right of insert_at
    by one column. Skips text inside double quotes."""
    def repl(match: re.Match) -> str:
        prefix_dollar = match.group(1) or ""
        letters = match.group(2)
        row_dollar = match.group(3) or ""
        row = match.group(4)
        return f"{prefix_dollar}{_shift_letter(letters, insert_at)}{row_dollar}{row}"

    parts = re.split(r'("[^"]*")', formula)
    for i in range(0, len(parts), 2):
        parts[i] = re.sub(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", repl, parts[i])
    return "".join(parts)


def _shifted_range(min_col: int, max_col: int, insert_at: int) -> tuple[int, int]:
    """Apply the insert to a (min_col, max_col) range.
    - entirely before insert → unchanged
    - entirely at/after      → both shifted +1
    - straddling             → min unchanged, max +1 (range absorbs the new col)
    """
    new_min = min_col if min_col < insert_at else min_col + 1
    new_max = max_col + 1 if max_col >= insert_at else max_col
    return new_min, new_max


def main() -> None:
    if not os.path.exists(TEMPLATE):
        sys.exit(f"template not found: {TEMPLATE}")
    if not os.path.exists(BACKUP):
        shutil.copy(TEMPLATE, BACKUP)
        print(f"backed up → {BACKUP}")

    wb = load_workbook(TEMPLATE, data_only=False)
    ws = wb["Property Tax"]

    # ----- 1) Snapshot every row-2 group merge BEFORE the insert.
    # Each entry: (anchor_label, old_min_col, old_max_col).
    group_merges: list[tuple[str, int, int]] = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row != 2 or rng.max_row != 2:
            continue
        label = ws.cell(2, rng.min_col).value
        group_merges.append((label, rng.min_col, rng.max_col))

    # ----- 2) Unmerge every row-2 group merge so the insert_cols doesn't
    # leave dangling references. We'll re-establish each at its shifted
    # coords AFTER the insert.
    for _, min_col, max_col in group_merges:
        ws.unmerge_cells(
            start_row=2, end_row=2,
            start_column=min_col, end_column=max_col,
        )

    # ----- 3) Capture an S4-style cell for cream-fill propagation.
    src_for_style = ws.cell(row=4, column=column_index_from_string("S"))
    saved_font = copy(src_for_style.font)
    saved_alignment = copy(src_for_style.alignment)
    saved_border = copy(src_for_style.border)

    # ----- 4) The insert itself.
    ws.insert_cols(INSERT_AT)
    print(f"inserted blank column at T (idx {INSERT_AT})")

    # ----- 5) Rewrite formulas that reference shifted columns.
    rewrites = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f" and isinstance(cell.value, str):
                new = _rewrite_formula(cell.value, INSERT_AT)
                if new != cell.value:
                    cell.value = new
                    rewrites += 1
    print(f"rewrote {rewrites} formula cells")

    # ----- 6) Re-establish row-2 group merges at their shifted/extended
    # coords. Place the label at the new anchor BEFORE merging so the
    # merge inherits it.
    for label, old_min, old_max in group_merges:
        new_min, new_max = _shifted_range(old_min, old_max, INSERT_AT)
        anchor = ws.cell(row=2, column=new_min)
        anchor.value = label
        if new_max > new_min:
            ws.merge_cells(
                start_row=2, end_row=2,
                start_column=new_min, end_column=new_max,
            )
        old_range = f"{get_column_letter(old_min)}2:{get_column_letter(old_max)}2"
        new_range = f"{get_column_letter(new_min)}2:{get_column_letter(new_max)}2"
        print(f"  group merge: {old_range} → {new_range}  ({label!r})")

    # ----- 7) Stamp the Confidence header in T3, matching S3's styling.
    src_header = ws.cell(row=3, column=column_index_from_string("S"))
    new_header = ws.cell(row=3, column=INSERT_AT)
    new_header.value = CONFIDENCE_HEADER
    new_header.font = copy(src_header.font)
    new_header.fill = copy(src_header.fill)
    new_header.alignment = copy(src_header.alignment)
    new_header.border = copy(src_header.border)

    # ----- 8) Cream fill on data rows.
    for r in range(4, ws.max_row + 1):
        c = ws.cell(row=r, column=INSERT_AT)
        c.fill = PatternFill(fill_type="solid", fgColor=CREAM)
        c.font = copy(saved_font)
        c.alignment = copy(saved_alignment)
        c.border = copy(saved_border)

    # ----- 9) Sample value + sensible width.
    ws.cell(row=4, column=INSERT_AT).value = "HIGH"
    ws.column_dimensions[get_column_letter(INSERT_AT)].width = 12

    wb.save(TEMPLATE)
    print(f"saved → {TEMPLATE}")


if __name__ == "__main__":
    main()
